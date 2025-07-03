import pandas as pd
import pyperclip
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import pyperclip
import os
import sys
import webbrowser
import tempfile
import subprocess
import platform
import re
import traceback

class RadiatorApp:
    def __init__(self, root):
        """
        Инициализация главного окна программы
        """
        self.root = root
        self.tooltip = None
        if not self.root.winfo_exists():  
            self.root = tk.Tk()  
        
        # Устанавливаем заголовок окна
        self.root.title("RadiaTool v1.9")
        
        # Получаем размеры экрана
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # Вычисляем координаты для центрирования по горизонтали и позиционирования у верхнего края
        x = (screen_width + 300) // 2  # 800 - примерная начальная ширина окна
        y = 0  # Позиция у верхнего края
        
        # Устанавливаем начальное положение окна
        self.root.geometry(f"+{x}+{y}")

        # По умолчанию подсказки включены
        self.show_tooltips_var = tk.BooleanVar(value=False) 

        # Устанавливаем иконку программы
        self.set_window_icon()
        
        # Настройка стилей
        self.setup_styles()
        
        # Инициализация переменных
        self.connection_var = tk.StringVar(value="VK-правое")
        self.radiator_type_var = tk.StringVar(value="10")
        self.bracket_var = tk.StringVar(value="Настенные кронштейны")
        self.radiator_discount_var = tk.StringVar(value="0")
        self.bracket_discount_var = tk.StringVar(value="0")
        self.entry_values = {}
        self.entries = {}
        
        # Загрузка данных
        self.load_data()
        
        # Создание интерфейса
        self.create_interface()
        self.update_radiator_types()
        self.show_selected_matrix()
        
        # Настройка размеров окна
        self.adjust_window_size()
        self.tooltip = None 

        # Обработчик закрытия окна
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        
        self._preview_window = None  # Добавляем атрибут для хранения ссылки на окно предпросмотра
        
        try:
            icon_path = self.resource_path("icon.ico")  
            self.root.iconbitmap(icon_path)
        except Exception as e:
            print(f"Не удалось установить иконку: {e}")

    def on_close(self):
        """Обработчик закрытия окна - уничтожает все подсказки"""
        if hasattr(self, 'vk_right_tooltip') and self.vk_right_tooltip:
            self.vk_right_tooltip.destroy()
        if hasattr(self, 'vk_left_tooltip') and self.vk_left_tooltip:
            self.vk_left_tooltip.destroy()
        self.root.destroy()

    def has_any_value(self):
        """Проверяет, есть ли хотя бы одно значение во всех матрицах"""
        return any(value for value in self.entry_values.values()) 
    
    def global_highlight(self):
        """Принудительно включает подсветку для всех матриц"""
        color = '#e6f3ff' if self.has_any_value() else 'white'
        for entry in self.entries.values():
            try:
                if entry.winfo_exists():
                    # Не перекрашиваем заполненные ячейки (они желтые)
                    if not entry.get().strip():
                        entry.config(bg=color)
            except tk.TclError:
                continue
    
    def create_image_tooltip(self, parent, image_path):
        """Создает подсказку с картинкой"""
        try:
            # Проверяем существование файла
            if not os.path.exists(image_path):
                print(f"Файл изображения не найден: {image_path}")
                return None
                
            tooltip = tk.Toplevel(parent)
            tooltip.wm_overrideredirect(True)
            tooltip.withdraw()
            
            img = tk.PhotoImage(file=image_path)
            label = ttk.Label(tooltip, image=img)
            label.image = img  # Сохраняем ссылку на изображение
            label.pack()
            
            return tooltip
        except Exception as e:
            print(f"Ошибка загрузки изображения для подсказки: {e}")
            return None

    def show_image_tooltip(self, tooltip, widget):
        """Показывает подсказку с картинкой под виджетом, центрируя по горизонтали относительно окна"""
        if tooltip:
            # Получаем координаты виджета для вертикального позиционирования
            y = widget.winfo_rooty() + widget.winfo_height() + 5  # 5px отступ под виджетом
            
            # Получаем координаты окна для горизонтального центрирования
            window_x = self.root.winfo_rootx()
            window_width = self.root.winfo_width()
            
            # Центрируем подсказку по горизонтали относительно окна
            tooltip_width = tooltip.winfo_reqwidth()
            x = window_x + (window_width - tooltip_width) // 2
            
            tooltip.wm_geometry(f"+{x}+{y}")
            tooltip.deiconify()

    def hide_image_tooltip(self, tooltip):
        """Скрывает подсказку с картинкой"""
        if tooltip and tooltip.winfo_exists():
            tooltip.withdraw()
    
    def open_file_default_app(self, path):
        """
        Открывает файл в программе, связанной с его типом в системе.
        Работает на Windows, macOS и Linux.
        """
        try:
            if platform.system() == "Windows":
                os.startfile(path)
            elif platform.system() == "Darwin":  # macOS
                subprocess.call(["open", path])
            else:  # Linux и другие
                subprocess.call(["xdg-open", path])
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть файл:\n{str(e)}")

    def hide_header_tooltip(self):
        """Скрывает подсказку"""
        if self.tooltip:
            self.tooltip.destroy()
            self.tooltip = None  

    def resource_path(self, relative_path):
        """Get absolute path to resource, works for dev and for PyInstaller"""
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)         

    def set_window_icon(self):
        try:
            icon_path = self.resource_path("favicon.ico")  
            self.root.iconbitmap(icon_path)
        except Exception as error:
            print(f"Не удалось установить иконку: {error}")

    def copy_articul_column(self, spec_data):
        """Копирование только столбца 'Артикул' (без итоговой строки)"""
        try:
            filtered_data = spec_data[spec_data["№"] != "Итого"]
            articuls = filtered_data["Артикул"].astype(str)
            cleaned_articuls = articuls.str.strip().replace('nan', '')
            pyperclip.copy('\n'.join(cleaned_articuls))
            
            # Переводим окно предпросмотра на задний план и активируем главное окно
            if hasattr(self, '_preview_window') and self._preview_window.winfo_exists():
                self._preview_window.lower()
                self.root.attributes('-topmost', True)
                self.root.focus_force()
                self.root.attributes('-topmost', False)
            
        except Exception as e:
            print(f"Ошибка копирования артикулов: {str(e)}")

    def copy_quantity_column(self, spec_data):
        """Копирование только столбца 'Кол-во' (без итоговой строки)"""
        try:
            filtered_data = spec_data[spec_data["№"] != "Итого"]
            quantities = filtered_data["Кол-во"].astype(str)
            cleaned_quantities = quantities.str.strip().replace('nan', '')
            pyperclip.copy('\n'.join(cleaned_quantities))
            
            # Переводим окно предпросмотра на задний план и активируем главное окно
            if hasattr(self, '_preview_window') and self._preview_window.winfo_exists():
                self._preview_window.lower()
                self.root.attributes('-topmost', True)
                self.root.focus_force()
                self.root.attributes('-topmost', False)
            
        except Exception as e:
            print(f"Ошибка копирования количества: {str(e)}")          

    def setup_styles(self):
        style = ttk.Style()
        style.theme_use('clam')

        # Задание глобального фона и основного цвета текста
        style.configure('.', background="#e1dfdf", foreground='#444141')  # Цвет фона и текста всего приложения

        # Фон Frame-контейнеров
        style.configure('TFrame', background='#dedede')  # Серый оттенок для фонов фреймов

        # Фон и шрифт Label-элементов
        style.configure('TLabel', background='#dedede', font=('Segoe UI', 9))  # Светло-серый фон и размер шрифта 9 пунктов

        # Внешний вид кнопок Button
        style.configure('TButton', font=('Segoe UI', 9), background='#955b67', foreground='white')  # Темно-красный фон и белый текст кнопки

        # Поля ввода Entry
        style.configure('TEntry', font=('Segoe UI', 9), fieldbackground='white')  # Белый фон для текстового поля ввода

        # Заголовочные LabelFrame
        style.configure('TLabelFrame', background="#dedede", font=('Segoe UI', 9, 'bold'))  # Фон светло-серый, жирный шрифт для подписей групп

        # Радио-кнопки
        style.configure('TRadiobutton', background='#dedede', font=('Segoe UI', 9))  # Аналогично TLabel, светлый серый фон

        # Меню-кнопки Menubutton
        style.configure('TMenubutton', font=('Segoe UI', 9), background='#955b67', foreground='white')  # Тёмно-красный фон и белый текст для выпадающих меню

        # Новый стиль для подсветки активной клетки
        style.configure('Highlight.TEntry', fieldbackground='#FFD700', font=('Segoe UI', 9))  # Жёлтый фон для выделенного элемента
        
        # В методе setup_styles добавьте новый стиль для кнопки:
        style.configure('Update.TButton', 
                    font=('Segoe UI', 9), 
                    background='#955b67', 
                    foreground='#444141')  # Устанавливаем цвет текста #444141

        # Изменение внешнего вида активных кнопок
        style.map('TButton',
                background=[('active', "#263168"), ('!disabled', "#7E1A2F")],
                foreground=[('!disabled', 'white')]
        )

        # Настройка стиля для Menubutton
        style.configure('TMenubutton', font=('Segoe UI', 9), background='#955b67', foreground='white')
        style.map('TMenubutton',
                background=[('active', '#263168'), ('!active', '#7E1A2F')],
                foreground=[('!active', 'white')]
        )
        
        # Стиль для меню (стрелки)
        style.configure('TMenubutton', 
                    background='#955b67', 
                    foreground='white',
                    arrowcolor='white')  # Устанавливаем цвет стрелки
        
        style.map('TMenubutton',
                background=[('active', '#263168'), ('!active', '#7E1A2F')],
                foreground=[('!active', 'white')],
                arrowcolor=[('!active', 'white')])  # Цвет стрелки в разных состояниях
        
    def copy_column(self, spec_data, column_name):
        """Копирует данные столбца в буфер обмена, исключая итоговую строку"""
        try:
            # Фильтрация данных (исключаем строку "Итого")
            filtered_data = spec_data[spec_data["№"] != "Итого"]
            
            # Проверка наличия столбца
            if column_name not in filtered_data.columns:
                raise ValueError(f"Столбец {column_name} не найден")
                
            # Преобразование данных в строки
            data_to_copy = filtered_data[column_name].astype(str)
            
            # Удаление пробелов и пустых значений
            cleaned_data = data_to_copy.str.strip().replace('nan', '')
            
            # Копирование в буфер через переносы строк
            pyperclip.copy('\n'.join(cleaned_data))
            
        except KeyError as ke:
            print(f"Ошибка ключа: {str(ke)}")
        except ValueError as ve:
            print(f"Ошибка значения: {str(ve)}")
        except Exception as e:
            print(f"Общая ошибка копирования: {str(e)}")

    def calculate_total_power(self, spec_data):
        """Рассчитывает суммарную мощность (Вт) с учетом количества"""
        total_power = 0.0
        
        for index, row in spec_data.iterrows():
            # Пропуск итоговой строки
            if row["№"] == "Итого":
                continue
                
            # Получение значений из строки
            power_str = str(row["Мощность, Вт"]).strip()
            qty = row["Кол-во"]
            
            try:
                # Конвертация мощности в число
                power = float(power_str) if power_str not in ['', 'nan', 'None'] else 0.0
                
                # Расчет и суммирование
                if power >= 0 and qty >= 0:
                    total_power += power * qty
                else:
                    print(f"Некорректные значения в строке {index}: мощность={power}, количество={qty}")
                    
            except ValueError:
                print(f"Ошибка конвертации мощности в строке {index}: '{power_str}'")
            except TypeError:
                print(f"Неправильный тип данных в строке {index}")
        
        return round(total_power, 2)  # Округление до 2 знаков

    def load_data(self):
        """
        Загружает данные из Excel-файлов, встроенных в EXE.
        Обрабатывает матрицу радиаторов и кронштейны.
        """
        try:
            # Получаем путь к файлу "Матрица.xlsx" внутри EXE
            self.file_path = self.resource_path("Матрица.xlsx") 
            
            # Проверяем существование файла
            if not os.path.exists(self.file_path):
                raise FileNotFoundError(f"Файл не найден: {self.file_path}")
            
            # Загружаем данные из Excel
            self.sheets = pd.read_excel(self.file_path, sheet_name=None, engine='openpyxl')
            
            # Обрабатываем лист с кронштейнами
            if "Кронштейны" in self.sheets:
                self.brackets_df = self.sheets["Кронштейны"].copy()
                self.brackets_df['Артикул'] = self.brackets_df['Артикул'].astype(str).str.strip()
                del self.sheets["Кронштейны"]
            else:
                self.brackets_df = pd.DataFrame()
            
            # Обрабатываем остальные листы
            for sheet_name, data in self.sheets.items():
                data['Артикул'] = data['Артикул'].astype(str).str.strip()
                data['Вес, кг'] = pd.to_numeric(data['Вес, кг'], errors='coerce').fillna(0)
                data['Объем, м3'] = pd.to_numeric(data['Объем, м3'], errors='coerce').fillna(0)
                data['Мощность, Вт'] = data.get('Мощность, Вт', '')
        except Exception as e:
            # Если произошла ошибка, показываем сообщение и закрываем программу
            messagebox.showerror("Ошибка", f"Ошибка загрузки данных: {str(e)}")
            self.root.destroy()

    def calculate_max_matrix_width(self):
        """
        Рассчитывает наибольший возможный размер матрицы среди доступных конфигураций.
        """
        max_width = 0
        
        # Возможные типы подключения
        connections = ["VK-правое", "VK-левое", "K-боковое"]
        
        # Доступные типы радиаторов для разных подключений
        types_for_connections = {
            "VK-правое": ["10", "11", "20", "21", "22", "30", "33"],
            "VK-левое": ["10", "11", "30", "33"],
            "K-боковое": ["10", "11", "20", "21", "22", "30", "33"]
        }
        
        # Проходим по каждому типу подключения и соответствующему списку типов радиаторов
        for conn in connections:
            available_types = types_for_connections.get(conn)
            
            # Подсчет длины матрицы
            for rad_type in available_types:
                sheet_name = f"{conn} {rad_type}"
                
                # Размеры матрицы зависят от количества колонок и строк
                num_columns = len(self.sheets[sheet_name].columns)
                num_rows = len(self.sheets[sheet_name])
                
                # Предположим, каждая ячейка имеет фиксированную ширину ~30 пикселей
                matrix_width = num_columns * 30 
                
                # Берем максимальное значение
                max_width = max(max_width, matrix_width)

        return max_width

    def adjust_window_size(self):
        # Обновляем геометрию, чтобы получить актуальные размеры виджетов
        self.root.update_idletasks()
        
        # 1. Рассчитываем ширину содержимого
        # Берем ширину матрицы или минимальную ширину 800
        matrix_width = self.scrollable_matrix_frame.winfo_reqwidth() if hasattr(self, 'scrollable_matrix_frame') else 800
        controls_width = self.top_panel.winfo_reqwidth() if hasattr(self, 'top_panel') else 800
        
        # Ширина окна = максимальная из ширины матрицы и управляющих элементов
        content_width = max(matrix_width, controls_width + 30)
        
        # 2. Рассчитываем высоту содержимого
        content_height = 0
        elements = [
            'top_panel', 'connection_frame', 'radiator_frame',
            'scrollable_matrix_frame', 'bracket_frame',
            'discount_frame', 'bottom_panel'
        ]
        
        for element in elements:
            if hasattr(self, element):
                widget = getattr(self, element)
                try:
                    content_height += widget.winfo_reqheight()
                except tk.TclError:
                    continue
        
        # 3. Добавляем отступы
        total_width = content_width + 20  # 10px с каждой стороны
        total_height = content_height + 30  # 15px сверху и снизу
        
        # 4. Ограничиваем размеры экраном
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        total_width = min(total_width, screen_width)
        total_height = min(total_height, screen_height)
        
        # 5. Устанавливаем размеры
        self.root.geometry(f"{total_width}x{total_height}")
        self.root.minsize(total_width, total_height + 50)

    def calculate_matrix_width(self):
        """Рассчитывает ширину матрицы радиаторов"""
        if not hasattr(self, 'scrollable_matrix_frame'):
            return 800  # Значение по умолчанию
            
        # Ширина матрицы = ширина всех столбцов + заголовки
        matrix_width = self.scrollable_matrix_frame.winfo_reqwidth()
        return max(matrix_width, 800)  # Не меньше 800px

    def calculate_controls_width(self):
        """Рассчитывает ширину управляющих элементов"""
        controls_width = 0
        if hasattr(self, 'top_panel'):
            controls_width = max(controls_width, self.top_panel.winfo_reqwidth())
        if hasattr(self, 'bottom_panel'):
            controls_width = max(controls_width, self.bottom_panel.winfo_reqwidth())
        
        return max(controls_width, 800)  # Не меньше 800px

    def get_brackets_list(self):
        """Возвращает список кронштейнов в формате для комбобокса"""
        brackets_list = []
        if not self.brackets_df.empty:
            for _, row in self.brackets_df.iterrows():
                brackets_list.append({
                    'Артикул': str(row['Артикул']).strip(),
                    'Наименование': str(row['Наименование']).strip()
                })
        return brackets_list        

    def refresh_matrix(self):
        """Полностью пересоздает матрицу с текущими значениями"""
        # Удаляем старые виджеты матрицы
        for widget in self.scrollable_matrix_frame.winfo_children():
            widget.destroy()
        
        # Очищаем текущее состояние
        self.entries.clear()

        # Получаем текущие параметры подключения и типа радиатора
        sheet_name = f"{self.connection_var.get()} {self.radiator_type_var.get()}"

        if sheet_name not in self.sheets:
            messagebox.showerror("Ошибка", f"Лист '{sheet_name}' не найден")
            return

        data = self.sheets[sheet_name]
        lengths = list(range(400, 2100, 100))
        heights = [300, 400, 500, 600, 900]

        # Создаем стиль для заголовков без рамок
        style = ttk.Style()
        style.configure('NoBorder.TLabel', relief='flat', borderwidth=0)

        # Заголовки столбцов (высоты)
        for j, h in enumerate(heights):
            label = ttk.Label(
                self.scrollable_matrix_frame, 
                text=str(h),
                width=8,
                style='NoBorder.TLabel',
                anchor="center"
            )
            label.grid(row=1, column=j+1, sticky="nsew")
        
        # Заголовки строк (длины)
        for i, l in enumerate(lengths):
            label = ttk.Label(
                self.scrollable_matrix_frame, 
                text=str(l),
                width=8,
                style='NoBorder.TLabel',
                anchor="center"
            )
            label.grid(row=i+2, column=0, sticky="nsew")
            
            # Ячейки с радиаторами
            for j, h in enumerate(heights):
                self.create_cell(sheet_name, data, l, h, i+2, j+1)

        # Подсвечиваем заполненные ячейки
        self.highlight_filled_cells()

    def create_interface(self):
        """
        Создает полный графический интерфейс приложения.
        Каждый элемент создается последовательно с комментариями.
        """
        # 1. Создаем главный контейнер (основной фрейм)
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill="both", expand=True, padx=5, pady=2)

        # 2. Создаем систему прокрутки для всего интерфейса
        canvas = tk.Canvas(main_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        # Настраиваем прокрутку
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        # Размещаем элементы прокрутки
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # 3. Создаем верхнюю панель (меню и логотип)
        self.top_panel = ttk.Frame(scrollable_frame)
        self.top_panel.pack(fill="x", pady=(0, 10))

        # Контейнер для кнопок меню с равномерным распределением
        menu_frame = ttk.Frame(self.top_panel)
        menu_frame.pack(fill="x", expand=True)

        buttons = [
            ("Создать", [
                ("Спецификацию METEOR", lambda: self.generate_spec("excel")),
                ("Файл METEOR CSV", lambda: self.generate_spec("csv"))
            ]),
            ("Загрузить из", [
                ("Спецификации METEOR", self.load_excel_spec),
                ("Файла METEOR CSV", self.load_csv_spec),
                ("Иной спецификации (в разработке!)", self.load_foreign_spec) 
            ]),
            ("Информация", [
                ("Лицензионное соглашение", lambda: self.show_info("agreement")),
                ("Инструкция по использованию программы", self.open_instruction_pdf),
                ("Прайс-лист от 19.03.25", self.open_price_list),
                ("Формуляр для регистрации проектов", self.open_project_form),
                ("Каталог отопительного оборудования", lambda: webbrowser.open("https://laggartt.ru/catalogs/laggartt/#p=386")),
                ("Паспорт на радиатор", lambda: webbrowser.open("https://b24.ez.meteor.ru/~yKwOz")),
                ("Сертификат соответствия", lambda: webbrowser.open("https://b24.ez.meteor.ru/~jmXyC")),
                ("Экспертное заключение (гигиена)", lambda: webbrowser.open("https://b24.ez.meteor.ru/~BGqta")),
                ("Расчёт мощностей радиаторов METEOR", self.open_power_calculation),
            ])
        ]

        for i, (btn_text, menu_items) in enumerate(buttons):
            btn = ttk.Menubutton(menu_frame, text=btn_text)
            menu = tk.Menu(btn, tearoff=0)

            if btn_text == "Информация":
                for idx, (item_text, command) in enumerate(menu_items):
                    menu.add_command(label=item_text, command=command)
                
                    if idx == 1 or idx == 4:
                        menu.add_separator()
            else:
                for item_text, command in menu_items:
                    menu.add_command(label=item_text, command=command)

            btn["menu"] = menu
            btn.grid(row=0, column=i, sticky="ew", padx=5)

        # Настраиваем равномерное распределение колонок
        for i in range(len(buttons)):
            menu_frame.grid_columnconfigure(i, weight=1)

        # 4. Создаем блок выбора типа подключения
        self.connection_frame = self.create_connection_frame(scrollable_frame)
        
        # 5. Создаем блок выбора типа радиатора
        self.radiator_frame = ttk.LabelFrame(scrollable_frame, text="Тип радиатора")
        self.radiator_frame.pack(fill="x", padx=5, pady=2)

        # Создаем Canvas для горизонтального текста "высота радиаторов, мм"
        self.height_label_canvas = tk.Canvas(
            scrollable_frame,
            width=200,
            height=30,
            highlightthickness=0,
            bg="#dedede"
        )
        self.height_label_canvas.pack(fill="x", pady=0)

        # Добавляем горизонтальный текст на Canvas
        self.height_label_canvas.create_text(
            250, 15,  # координаты центра Canvas
            text="высота радиаторов, мм",
            font=("Segoe UI", 9),
            fill="#444141"
        )

        # 6. Создаем контейнер для матрицы радиаторов с вертикальным текстом слева
        self.matrix_container = ttk.Frame(scrollable_frame)
        self.matrix_container.pack(fill="both", expand=True, padx=2, pady=2)

        # Создаем Canvas для вертикального текста слева от матрицы
        self.length_label_canvas = tk.Canvas(
            self.matrix_container, 
            width=15, 
            height=200, 
            highlightthickness=0, 
            bg="#dedede"
        )
        self.length_label_canvas.grid(row=0, column=0, sticky="ns", padx=(0, 0))  # Колонка 0, с отступом справа

        # Добавляем вертикальный текст на Canvas (центрируем по ширине и высоте)
        self.length_label_canvas.create_text(
            8, 250,  # координаты центра Canvas 
            text="длина радиаторов, мм",
            angle=90,
            font=("Segoe UI", 9),
            fill="#444141"
        )

        # Внутренний фрейм матрицы радиаторов с прокруткой
        self.scrollable_matrix_frame = ttk.Frame(self.matrix_container, style='TFrame')
        self.scrollable_matrix_frame.grid(row=0, column=1, sticky="nsew")  # Сдвинули на колонку 1

        # Настраиваем растяжение колонок grid в matrix_container
        self.matrix_container.columnconfigure(0, weight=0)  # Текст фиксирован
        self.matrix_container.columnconfigure(1, weight=1)  # Матрица растягивается
        self.matrix_container.rowconfigure(0, weight=1)

        # 7. Создаем блок выбора кронштейнов
        self.bracket_frame = ttk.Frame(scrollable_frame)
        self.bracket_frame.pack(fill="x", padx=5, pady=2)

        # Левая часть - выбор типа кронштейнов
        radio_frame = ttk.Frame(self.bracket_frame)
        radio_frame.pack(side="left", fill="y", expand=False)

        # Варианты кронштейнов
        brackets = ["Настенные кронштейны", "Напольные кронштейны", "Без кронштейнов"]
        for bracket in brackets:
            ttk.Radiobutton(
                radio_frame,
                text=bracket,
                variable=self.bracket_var,
                value=bracket
            ).pack(anchor="w", padx=10, pady=2)

        # Правая часть - чекбокс для показа характеристик и кнопка обновления
        right_frame = ttk.Frame(self.bracket_frame)
        right_frame.pack(side="right", fill="y", padx=(20, 0))

        # Упаковка элементов с выравниванием по правому краю
        ttk.Checkbutton(
            right_frame,
            text="Показывать\nпараметры",
            variable=self.show_tooltips_var,
            command=self.toggle_tooltips,
            style="WrapText.TCheckbutton"
        ).pack(anchor="e", pady=2)  

        # Кнопка проверки обновлений
        ttk.Button(
            right_frame,
            text="Проверить\nобновление",
            command=lambda: webbrowser.open("https://b24.engpx.ru/~HinAV"),
            style="Neutral.TButton"
        ).pack(anchor="e", pady=1)  

        # Стиль для переноса текста в чекбоксе
        style = ttk.Style()
        style.configure("WrapText.TCheckbutton", wraplength=100)
        style.configure("Neutral.TButton", 
                    background="#dedede",  # Цвет фона как у основного интерфейса
                    foreground="#444141",  # Цвет текста
                    relief="solid",        # Граница
                    borderwidth=1)         # Толщина границы

        # 8. Создаем блок скидок
        self.discount_frame = ttk.Frame(scrollable_frame)
        self.discount_frame.pack(fill="x", padx=5, pady=2)

        # Скидка на радиаторы
        ttk.Label(self.discount_frame, 
                text="Скидка на радиаторы, %:", 
                width=25).pack(side="left")
        ttk.Entry(
            self.discount_frame,
            textvariable=self.radiator_discount_var,
            width=5,
            validate="key",
            validatecommand=(self.discount_frame.register(self.validate_discount), '%P')
        ).pack(side="left", padx=2)

        # Скидка на кронштейны
        ttk.Label(self.discount_frame, 
                text="кронштейны, %:", 
                width=15).pack(side="left")
        ttk.Entry(
            self.discount_frame,
            textvariable=self.bracket_discount_var,
            width=5,
            validate="key",
            validatecommand=(self.discount_frame.register(self.validate_discount), '%P')
        ).pack(side="left", padx=2)

        # 9. Создаем нижнюю панель с кнопками
        self.bottom_panel = ttk.Frame(scrollable_frame)
        self.bottom_panel.pack(fill="x", pady=10)

        # Кнопка предпросмотра
        self.preview_btn = ttk.Button(
            self.bottom_panel,
            text="Предпросмотр",
            command=self.preview_spec,
            width=15
        )
        self.preview_btn.pack(side="left", padx=5)

        # Вместо кнопки "Сброс" теперь размещаем логотип
        try:
            logo_path = self.resource_path("Lagar.png")
            logo_img = tk.PhotoImage(file=logo_path)
            self.logo_label = ttk.Label(self.bottom_panel, image=logo_img)
            self.logo_label.image = logo_img  # Сохраняем ссылку на изображение
            self.logo_label.pack(side="left", padx=10)  # Слева от кнопок
            # Делаем логотип кликабельным
            self.logo_label.bind("<Button-1>", 
                            lambda e: webbrowser.open("https://laggartt.ru"))
            self.logo_label.bind("<Enter>", 
                            lambda e: self.logo_label.config(relief="raised"))
            self.logo_label.bind("<Leave>", 
                            lambda e: self.logo_label.config(relief="flat"))
        except Exception as e:
            print(f"Не удалось загрузить логотип: {e}")

        # Кнопка сброса 
        self.reset_btn = ttk.Button(
            self.bottom_panel,
            text="Сброс",
            command=self.reset_fields,
            width=15
        )
        self.reset_btn.pack(side="right", padx=5)

        # 10. Обновляем размеры окна
        self.adjust_window_size()
        self.tooltip = None

    def open_instruction_pdf(self):
        """Открывает файл инструкции PDF"""
        try:
            # Получаем путь к файлу внутри EXE
            pdf_path = self.resource_path("inst.pdf")
            
            if os.path.exists(pdf_path):
                self.open_file_default_app(pdf_path)
            else:
                messagebox.showerror("Ошибка", f"Файл инструкции не найден по пути: {pdf_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть файл инструкции:\n{str(e)}")    

    def open_power_calculation(self):
        try:
            # Получаем путь к файлу внутри EXE
            power_calc_path = self.resource_path("Расчет мощностей METEOR.xlsx")
            
            if os.path.exists(power_calc_path):
                self.open_file_default_app(power_calc_path)
            else:
                messagebox.showerror("Ошибка", f"Файл не найден по пути: {power_calc_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка открытия файла: {str(e)}")

    def load_foreign_spec(self):
        """Загружает спецификацию от других производителей и подбирает аналоги Meteor"""
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel Files", "*.xlsx *.xls"), ("CSV Files", "*.csv")]
        )
        if not file_path:
            return

        try:
            # Определяем движок для чтения
            if file_path.endswith('.xlsx'):
                engine = 'openpyxl'
            elif file_path.endswith('.xls'):
                engine = 'xlrd'
            elif file_path.endswith('.csv'):
                return self.load_csv_spec()
            else:
                messagebox.showerror("Ошибка", "Неподдерживаемый формат файла")
                return

            # Читаем весь файл для анализа
            df = pd.read_excel(file_path, engine=engine, header=None)
            
            # 1. Поиск строки с данными (не заголовками)
            data_start_row = 0
            for i, row in df.iterrows():
                # Ищем первую строку, где есть данные о радиаторах
                if any('радиатор' in str(cell).lower() or 
                    'vk' in str(cell).lower() or 
                    'k' in str(cell).lower() or 
                    '/' in str(cell) for cell in row):
                    data_start_row = i
                    break
            
            # 2. Определяем столбцы с наименованием и количеством
            name_col = None
            qty_col = None
            
            # Проверяем возможные варианты названий столбцов
            for j in range(len(df.columns)):
                col_data = df.iloc[data_start_row:data_start_row+5, j].astype(str).str.lower()
                
                # Проверяем на наличие ключевых слов в столбце
                if any('наимен' in s or 'радиатор' in s or 'назван' in s or 'тип' in s for s in col_data):
                    name_col = j
                elif any('кол-во' in s or 'количеств' in s or 'qty' in s or 'шт' in s for s in col_data):
                    qty_col = j
            
            # Если не нашли стандартные названия - берем первые два столбца
            if name_col is None:
                name_col = 0
            if qty_col is None:
                qty_col = 1 if len(df.columns) > 1 else 0

            # 3. Извлекаем данные
            data_rows = []
            for i in range(data_start_row, len(df)):
                name = str(df.iloc[i, name_col]).strip() if name_col is not None else ""
                qty = df.iloc[i, qty_col] if qty_col is not None else 0
                
                # Пропускаем пустые строки и строки "Итого"
                if not name or name.lower() == 'итого':
                    continue
                    
                try:
                    qty = float(qty)
                    if qty > 0 and ('радиатор' in name.lower() or 
                                    'vk' in name.lower() or 
                                    'k' in name.lower() or 
                                    '/' in name):
                        data_rows.append((name, int(qty)))
                except (ValueError, TypeError):
                    continue

            # 4. Очищаем текущие значения
            self.entry_values.clear()
            total_loaded = 0
            total_qty = 0
            not_found = []
            unknown_format = []
            long_radiators = []
            similar_loaded = []
            
            # Создаем DataFrame для соответствия
            correspondence_df = pd.DataFrame(columns=[
                "Оригинальное наименование", "Количество", 
                "Аналог Meteor", "Артикул Meteor", "Комментарий"
            ])

            # 5. Обрабатываем каждую строку
            for name, qty in data_rows:
                # 5.1. Определяем тип подключения
                name_lower = name.lower()
                
                # Все радиаторы с VK в начале названия - VK-правое подключение
                if re.match(r'^vk\s*\d+', name_lower, re.IGNORECASE):
                    connection = "VK-правое"
                # Радиаторы с VK в названии - VK-правое подключение
                elif 'vk' in name_lower or 'нижн' in name_lower:
                    connection = "VK-правое"
                # Все остальные - K-боковое подключение
                else:
                    connection = "K-боковое"

                # 5.2. Извлекаем параметры из названия
                rad_type = None
                height = None
                length = None
                
                # Вариант 1: Формат типа "VK 11-400-400" или "VK 11/400/800"
                match = re.match(r'^[vk]\s*(\d+)\s*[\-\s\/]\s*(\d+)\s*[\-\s\/]\s*(\d+)', name, re.IGNORECASE)
                if match:
                    rad_type = match.group(1)
                    height = match.group(2)
                    length = match.group(3)
                else:
                    # Вариант 2: Формат типа "11\500\400" или "тип 11/500/400"
                    match = re.search(r'(тип\s*)?(\d+)[\\\/\s\-]*(\d+)[\\\/\s\-]*(\d+)', name, re.IGNORECASE)
                    if match:
                        rad_type = match.group(2)
                        height = match.group(3)
                        length = match.group(4)
                    else:
                        # Вариант 3: Формат типа "K-Profil 11 500 400"
                        match = re.search(r'(\d+)\s+(\d+)\s+(\d+)', name)
                        if match:
                            rad_type = match.group(1)
                            height = match.group(2)
                            length = match.group(3)
                        else:
                            # Вариант 4: Формат типа "тип 11 / 500 / 400"
                            match = re.search(r'тип\s*(\d+)\s*/\s*(\d+)\s*/\s*(\d+)', name, re.IGNORECASE)
                            if match:
                                rad_type = match.group(1)
                                height = match.group(2)
                                length = match.group(3)
                            else:
                                unknown_format.append(name)
                                correspondence_df.loc[len(correspondence_df)] = [
                                    name, qty, "", "", "Не распознан формат"
                                ]
                                continue

                if not rad_type or not height or not length:
                    unknown_format.append(name)
                    correspondence_df.loc[len(correspondence_df)] = [
                        name, qty, "", "", "Не удалось определить параметры"
                    ]
                    continue
                    
                # Корректируем тип радиатора
                rad_type = rad_type.strip()
                
                # Проверяем, что тип поддерживается
                supported_types = ["10", "11", "20", "21", "22", "30", "33"]
                if rad_type not in supported_types:
                    not_found.append(f"{name} (неподдерживаемый тип {rad_type})")
                    correspondence_df.loc[len(correspondence_df)] = [
                        name, qty, "", "", f"Неподдерживаемый тип {rad_type}"
                    ]
                    continue

                # Корректируем высоту
                try:
                    height = int(height)
                    if height not in [300, 400, 500, 600, 900]:
                        original_height = height
                        if height < 300:
                            height = 300
                        elif height < 400:
                            height = 400
                        elif height < 500:
                            height = 500
                        elif height < 600:
                            height = 600
                        else:
                            height = 900
                        similar_loaded.append(f"{name} → высота {height} (была {original_height})")
                    
                    length = int(length)
                    if length > 2000:
                        long_radiators.append(f"{name} (длина {length} мм)")
                        correspondence_df.loc[len(correspondence_df)] = [
                            name, qty, "", "", f"Длина {length} мм > 2000 мм"
                        ]
                        continue
                    elif length < 400:
                        original_length = length
                        length = 400
                        similar_loaded.append(f"{name} → длина {length} (была {original_length})")
                    else:
                        original_length = length
                        length = round(length / 100) * 100
                        if original_length != length:
                            similar_loaded.append(f"{name} → длина {length} (была {original_length})")
                except ValueError:
                    unknown_format.append(f"{name} (ошибка в размерах)")
                    correspondence_df.loc[len(correspondence_df)] = [
                        name, qty, "", "", "Ошибка в размерах"
                    ]
                    continue

                # Формируем артикул Meteor
                sheet_name = f"{connection} {rad_type}"
                
                if sheet_name not in self.sheets:
                    not_found.append(f"{name} (неподдерживаемый тип {rad_type})")
                    correspondence_df.loc[len(correspondence_df)] = [
                        name, qty, "", "", f"Неподдерживаемый тип {rad_type}"
                    ]
                    continue

                # Ищем радиатор с такой же высотой и длиной
                pattern = f"/{height}/{length}"
                meteor_data = self.sheets[sheet_name]
                match = meteor_data[meteor_data['Наименование'].str.contains(pattern, na=False)]
                
                if not match.empty:
                    product = match.iloc[0]
                    art = str(product['Артикул']).strip()
                    
                    if (sheet_name, art) in self.entry_values:
                        current_qty = self.parse_quantity(self.entry_values[(sheet_name, art)])
                        self.entry_values[(sheet_name, art)] = str(current_qty + qty)
                    else:
                        self.entry_values[(sheet_name, art)] = str(qty)
                    
                    correspondence_df.loc[len(correspondence_df)] = [
                        name, qty, 
                        product['Наименование'], art,
                        "Успешно загружен"
                    ]
                    
                    total_loaded += 1
                    total_qty += qty
                else:
                    possible_lengths = [l for l in range(400, 2100, 100)]
                    closest_length = min(possible_lengths, key=lambda x: abs(x - length))
                    
                    pattern = f"/{height}/{closest_length}"
                    match = meteor_data[meteor_data['Наименование'].str.contains(pattern, na=False)]
                    
                    if not match.empty:
                        product = match.iloc[0]
                        art = str(product['Артикул']).strip()
                        
                        if (sheet_name, art) in self.entry_values:
                            current_qty = self.parse_quantity(self.entry_values[(sheet_name, art)])
                            self.entry_values[(sheet_name, art)] = str(current_qty + qty)
                        else:
                            self.entry_values[(sheet_name, art)] = str(qty)
                        
                        similar_loaded.append(f"{name} → {product['Наименование']} (длина {length}→{closest_length} мм)")
                        correspondence_df.loc[len(correspondence_df)] = [
                            name, qty, 
                            product['Наименование'], art,
                            f"Длина скорректирована {length}→{closest_length} мм"
                        ]
                        total_loaded += 1
                        total_qty += qty
                    else:
                        not_found.append(f"{name} (не найден аналог Meteor {height}x{length})")
                        correspondence_df.loc[len(correspondence_df)] = [
                            name, qty, "", "", 
                            f"Не найден аналог Meteor {height}x{length}"
                        ]

            # 6. Обновляем интерфейс
            self.refresh_matrix()
            self.global_highlight() 
            
            # 7. Формируем сообщение о результате
            msg = f"Успешно загружено: {total_loaded} позиций\nОбщее количество: {total_qty}"
            
            if similar_loaded:
                msg += f"\n\nЗагружены аналоги для {len(similar_loaded)} позиций:"
                for item in similar_loaded[:5]:
                    msg += f"\n- {item}"
                if len(similar_loaded) > 5:
                    msg += "\n..."
            
            if long_radiators:
                msg += f"\n\nРадиаторы с длиной >2000 мм (не включены):\n" + "\n".join(long_radiators[:5])
                if len(long_radiators) > 5:
                    msg += "\n..."
            
            if not_found:
                msg += f"\n\nНе найдены аналоги для {len(not_found)} позиций"
                if len(not_found) <= 10:
                    msg += ":\n" + "\n".join(not_found[:10])
                else:
                    msg += " (первые 10):\n" + "\n".join(not_found[:10]) + "\n..."
            
            if unknown_format:
                msg += f"\n\nНе распознан формат для {len(unknown_format)} позиций"
                if len(unknown_format) <= 10:
                    msg += ":\n" + "\n".join(unknown_format[:10])
                else:
                    msg += " (первые 10):\n" + "\n".join(unknown_format[:10]) + "\n..."
            
            messagebox.showinfo("Результат загрузки", msg)
            
            if not correspondence_df.empty:
                self.show_correspondence_table(correspondence_df)
                if hasattr(self, '_correspondence_window') and self._correspondence_window.winfo_exists():
                    self._correspondence_window.attributes('-topmost', True)
                    self._correspondence_window.after(100, lambda: self._correspondence_window.attributes('-topmost', False))

            self._correspondence_df = correspondence_df

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить файл:\n{str(e)}")
            
    def show_correspondence_table(self, correspondence_df):
        """Показывает таблицу соответствия в отдельном окне"""
        if correspondence_df.empty:
            return
        
        # Создаем окно для таблицы соответствия
        self._correspondence_window = tk.Toplevel(self.root)
        self._correspondence_window.title("Таблица соответствия")
        self._correspondence_window.geometry("1200x600")
        
        
        # Главный контейнер
        main_frame = ttk.Frame(self._correspondence_window)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Фрейм для таблицы с прокруткой
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill="both", expand=True)
        
        # Создаем Treeview
        columns = list(correspondence_df.columns)
        tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show="headings",
            selectmode="extended"
        )
        
        # Настройка столбцов
        col_widths = {
            "Оригинальное наименование": 400,
            "Количество": 80,
            "Аналог Meteor": 400,
            "Артикул Meteor": 120,
            "Комментарий": 200
        }
        
        # Конфигурация заголовков
        for col in columns:
            tree.heading(col, text=col)
            # Для столбца "Количество" устанавливаем центрирование
            if col == "Количество":
                tree.column(col, width=col_widths.get(col, 100), anchor="center")
            else:
                tree.column(col, width=col_widths.get(col, 100), anchor="w")
        
        # Добавление данных
        for _, row in correspondence_df.iterrows():
            tree.insert("", "end", values=list(row))
        
        # Прокрутка
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Упаковка элементов
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # Фрейм для кнопок
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=10)
        
        # Кнопка сохранения
        ttk.Button(
            button_frame,
            text="Сохранить таблицу соответствия",
            command=lambda: self.save_correspondence_table(correspondence_df)
        ).pack(side="left", padx=5)
        
        # Кнопка закрытия
        ttk.Button(
            button_frame,
            text="Закрыть",
            command=self._close_correspondence_window
        ).pack(side="right", padx=5)

    def _close_correspondence_window(self):
        """Закрывает окно соответствия и очищает ссылку"""
        if hasattr(self, '_correspondence_window') and self._correspondence_window and self._correspondence_window.winfo_exists():
            self._correspondence_window.destroy()
        self._correspondence_window = None

    def save_correspondence_table(self, correspondence_df):
        """Сохраняет таблицу соответствия в файл"""
        # Устанавливаем имя файла по умолчанию
        default_filename = "Таблица соответствия переподбора на радиаторы METEOR"
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("CSV Files", "*.csv")],
            title="Сохранить таблицу соответствия",
            initialfile=default_filename  # Устанавливаем имя файла по умолчанию
        )
        
        if not file_path:
            return
        
        try:
            if file_path.endswith('.xlsx'):
                correspondence_df.to_excel(file_path, index=False)
            elif file_path.endswith('.csv'):
                correspondence_df.to_csv(file_path, index=False, sep=';', encoding='utf-8-sig')
            
            messagebox.showinfo("Успешно", f"Таблица соответствия сохранена:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить файл:\n{str(e)}")

    def create_connection_frame(self, parent):
        frame = ttk.LabelFrame(parent, text="Вид подключения")
        frame.pack(fill="x", padx=5, pady=2)

        # Создаем подсказки для каждого типа подключения
        self.vk_right_tooltip = self.create_image_tooltip(frame, self.resource_path("1.png"))
        self.vk_left_tooltip = self.create_image_tooltip(frame, self.resource_path("2.png"))
        self.k_side_tooltip = self.create_image_tooltip(frame, self.resource_path("3.png"))

        # Варианты подключения
        connections = [
            ("VK-нижнее\nправое", "VK-правое"),
            ("VK-нижнее\nлевое", "VK-левое"),
            ("K-боковое", "K-боковое")
        ]

        for text, value in connections:
            btn = ttk.Radiobutton(
                frame,
                text=text,
                variable=self.connection_var,
                value=value,
                command=self.update_radiator_types
            )
            btn.pack(side="left", padx=10, pady=2)
            
            # Добавляем подсказки для всех типов подключения
            if value == "VK-правое" and self.vk_right_tooltip:
                btn.bind("<Enter>", lambda e, btn=btn: self.show_image_tooltip(self.vk_right_tooltip, btn))
                btn.bind("<Leave>", lambda e: self.hide_image_tooltip(self.vk_right_tooltip))
            elif value == "VK-левое" and self.vk_left_tooltip:
                btn.bind("<Enter>", lambda e, btn=btn: self.show_image_tooltip(self.vk_left_tooltip, btn))
                btn.bind("<Leave>", lambda e: self.hide_image_tooltip(self.vk_left_tooltip))
            elif value == "K-боковое" and self.k_side_tooltip:
                btn.bind("<Enter>", lambda e, btn=btn: self.show_image_tooltip(self.k_side_tooltip, btn))
                btn.bind("<Leave>", lambda e: self.hide_image_tooltip(self.k_side_tooltip))

        return frame

    def create_radiator_frame(self, parent):
        self.radiator_frame = ttk.LabelFrame(parent, text="Тип радиатора")
        self.radiator_frame.pack(fill="x", padx=5, pady=2)

    def create_matrix_frame(self, parent):
        # Контейнер для всей матрицы
        self.matrix_container = ttk.Frame(parent)
        self.matrix_container.pack(fill="both", expand=True, padx=5, pady=2)

        # Внутренний фрейм с прокруткой
        self.scrollable_matrix_frame = ttk.Frame(self.matrix_container, style='TFrame')
        self.scrollable_matrix_frame.grid(row=0, column=0, sticky="nsew")

        # Зафиксируем возможность расширения контейнера
        self.matrix_container.rowconfigure(0, weight=1)
        self.matrix_container.columnconfigure(0, weight=1)


    def create_discount_frame(self, parent):
        frame = ttk.Frame(parent) 
        frame.pack(fill="x", padx=5, pady=2)

        # Поле ввода скидок на радиаторы
        ttk.Label(frame, text="Скидка на: радиаторы, %:", width=25).pack(side="left")
        ttk.Entry(
            frame,
            textvariable=self.radiator_discount_var,
            width=5,
            validate="key",
            validatecommand=(frame.register(self.validate_discount), '%P')
        ).pack(side="left", padx=2)

        # Поле ввода скидок на кронштейны
        ttk.Label(frame, text="кронштейны, %:", width=15).pack(side="left")
        ttk.Entry(
            frame,
            textvariable=self.bracket_discount_var,
            width=5,
            validate="key",
            validatecommand=(frame.register(self.validate_discount), '%P')
        ).pack(side="left", padx=2)

    def create_action_buttons(self, parent):
        frame = ttk.Frame(parent)
        frame.pack(fill="x", pady=10)
        
        ttk.Button(frame, text="Предпросмотр", command=self.preview_spec).pack(side="left", padx=5)
        ttk.Button(frame, text="Сброс", command=self.reset_fields).pack(side="left", padx=5)
        
        # Добавляем чекбокс для управления подсказками
        ttk.Checkbutton(
            frame, 
            text="Подсказки", 
            variable=self.show_tooltips_var,
            command=self.toggle_tooltips
        ).pack(side="left", padx=10)
        
        ttk.Button(frame, text="Справка", command=self.show_help).pack(side="right", padx=5)

    def toggle_tooltips(self):
        """Включает/выключает подсказки при наведении"""
        if self.show_tooltips_var.get():
            # Включаем подсказки - обработчики уже привязаны
            pass
        else:
            # Выключаем подсказки - скрываем текущую, если есть
            self.hide_tooltip_on_leave()

    def open_price_list(self):
        try:
            # Получаем путь к файлу внутри EXE
            price_list_path = self.resource_path("Прайс-лист.xlsx")
            
            if os.path.exists(price_list_path):
                os.startfile(price_list_path)
            else:
                messagebox.showerror("Ошибка", f"Файл не найден по пути: {price_list_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка открытия файла: {str(e)}")
            
    def open_project_form(self):
        try:
            project_form_path = self.resource_path("Формуляр для регистрации проектов.xlsm")
        
            if os.path.exists(project_form_path):
                os.startfile(project_form_path)
            else:
                messagebox.showerror("Ошибка", "Файл 'Формуляр для регистрации проектов.xlsm' не найден.")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка открытия файла: {str(e)}")

    def show_info(self, info_type):
        if info_type == "help":
            help_text = """
            ИНСТРУКЦИЯ ПО ИСПОЛЬЗОВАНИЮ ПРОГРАММЫ "ПОДБОР РАДИАТОРОВ METEOR"

            1. ОСНОВНЫЕ ВОЗМОЖНОСТИ:
            - Подбор радиаторов по типу, размерам и подключению
            - Автоматический расчет кронштейнов
            - Формирование спецификаций или CSV файлов
            - Импорт данных из существующих спецификаций (xlsx/CSV), при импорте программа автоматически суммирует одинаковые 
              артикулы. Поддерживается загрузка спецификаций других производителей (в разработке)

            2. ПОРЯДОК РАБОТЫ:

            2.1. Выбор параметров:
            - В левой части окна выберите:
            * Вид подключения (VK-правое, VK-левое, K-боковое)
            * Тип радиатора (10, 11, 20, 21, 22, 30, 33)
            * Тип крепления (настенные/напольные кронштейны или без крепления)
            * Скидки (при необходимости)

            2.2. Заполнение матрицы:
            - В центральной таблице укажите количество радиаторов для нужных размеров
            - Можно вводить значения вида "1+2" для сложения количеств
            - Для просмотра характеристик радиатора наведите курсор на ячейку, используйте чекбокс "Показывать параметры"

            2.3. Формирование спецификации:
            - Нажмите "Предпросмотр" для проверки данных
            - В окне предпросмотра можно:
            * Удалить ненужные позиции (правой кнопкой мыши)
            * Добавить дополнительные кронштейны
            * Скопировать артикулы или количества (клик по заголовку столбца)
            - Сохраните спецификацию в нужном формате (xlsx или CSV)

            3. ДОПОЛНИТЕЛЬНЫЕ ФУНКЦИИ:

            3.1. Информационные материалы:
            - В меню "Информация" доступны:
            * Прайс-лист
            * Формуляр для регистрации проектов
            * Техническая документация
            * Сертификаты

            4. ТЕХНИЧЕСКАЯ ПОДДЕРЖКА:
            При возникновении вопросов обращайтесь:
            - Электронная почта: mt@laggartt.ru
            """
            messagebox.showinfo("Инструкция", help_text)
        elif info_type == "agreement":
            agreement_text = """
            ЛИЦЕНЗИОННОЕ СОГЛАШЕНИЕ
            НА ИСПОЛЬЗОВАНИЕ ПРОГРАММНОГО ОБЕСПЕЧЕНИЯ "METEOR"
            
            Настоящее Лицензионное соглашение (далее – «Соглашение») регулирует условия использования программного обеспечения «METEOR» (далее – «ПО»), распространяемого на безвозмездной основе.
            
            1. Права на интеллектуальную собственность
            1.1. Программное обеспечение, включая его исходный код, является интеллектуальной собственностью ООО «Термотехника Энгельс» (далее – «Правообладатель»).
            1.2. Любое несанкционированное копирование, модификация или распространение ПО запрещено.
            
            Реквизиты Правообладателя:
            ООО «Термотехника Энгельс»
            Россия, Саратовская область, г. Энгельс,
            Проспект Ф. Энгельса, 139
            Официальный сайт: https://laggartt.ru
            
            2. Назначение ПО
            2.1. Программное обеспечение предназначено для формирования спецификаций на радиаторы Meteor в соответствии с программой поставки ООО «Термотехника Энгельс».
            
            3. Условия использования результатов работы ПО
            3.1. Результаты, полученные с использованием ПО, носят оценочный характер и не могут рассматриваться как точные данные.
            3.2. ПО предназначено исключительно для предварительной оценки оборудования и не может служить основанием для проектных решений без дополнительной проверки.
            3.3. Использование ПО не освобождает Пользователя от соблюдения действующих отраслевых стандартов (DIN, ГОСТ), а также территориальных норм и правил.
            3.4. Пользователь самостоятельно несет ответственность за:
            
            корректность подбора оборудования;
            
            соответствие материалов и компонентов требованиям проекта;
            
            последствия, вызванные некорректным использованием ПО.
            3.5. ООО «Термотехника Энгельс» не несет ответственности за ущерб, возникший в результате применения данных, полученных с использованием ПО.
            
            4. Условия распространения и использования ПО
            4.1. Пользователь вправе устанавливать и использовать ПО на любом количестве компьютеров, в том числе в локальной сети организации.
            4.2. Использование ПО разрешено несколькими сотрудниками при условии соблюдения настоящего Соглашения.
            
            5. Срок действия и изменения
            5.1. Действие настоящего Соглашения распространяется на версию ПО, актуальную на дату его выпуска.
            5.2. ООО «Термотехника Энгельс» оставляет за собой право вносить изменения в номенклатуру оборудования без дополнительного уведомления Пользователей.
            
            ООО «Термотехника Энгельс»
            (05.2025)
            """
            messagebox.showinfo("Лицензионное соглашение", agreement_text)        

    def update_radiator_types(self):
        for widget in self.radiator_frame.winfo_children():
            widget.destroy()

        container = ttk.Frame(self.radiator_frame)
        container.pack(fill="x")

        types = ["10", "11", "30", "33"] if self.connection_var.get() == "VK-левое" else ["10", "11", "20", "21", "22", "30", "33"]
        
        # Устанавливаем значение по умолчанию "10" при переключении типа подключения
        self.radiator_type_var.set("10")
        
        for t in types:
            ttk.Radiobutton(
                container,
                text=t,
                variable=self.radiator_type_var,
                value=t,
                command=self.show_selected_matrix
            ).pack(side="left", padx=5, pady=2)

        self.show_selected_matrix()

    def show_selected_matrix(self):
        for widget in self.scrollable_matrix_frame.winfo_children():
            widget.destroy()
        
        sheet_name = f"{self.connection_var.get()} {self.radiator_type_var.get()}"
        if sheet_name not in self.sheets:
            messagebox.showerror("Ошибка", f"Лист '{sheet_name}' не найден")
            return
        
        data = self.sheets[sheet_name]
        lengths = list(range(400, 2100, 100))
        heights = [300, 400, 500, 600, 900]

        # Создаем стиль для заголовков без рамок
        style = ttk.Style()
        style.configure('NoBorder.TLabel', relief='flat', borderwidth=0)

        # Заголовки столбцов (высоты) теперь идут в строке 1
        for j, h in enumerate(heights):
            label = ttk.Label(
                self.scrollable_matrix_frame, 
                text=str(h),
                width=8,
                style='NoBorder.TLabel',
                anchor="center"
            )
            label.grid(row=1, column=j+1, sticky="nsew")
        
        # Заголовки строк (длины) начинаются со строки 2
        for i, l in enumerate(lengths):
            label = ttk.Label(
                self.scrollable_matrix_frame, 
                text=str(l),
                width=8,
                style='NoBorder.TLabel',
                anchor="center"
            )
            label.grid(row=i+2, column=0, sticky="nsew")
            
            # Ячейки с радиаторами
            for j, h in enumerate(heights):
                self.create_cell(sheet_name, data, l, h, i+2, j+1)

        # Ограничиваем максимальную ширину матрицы
        max_matrix_width = 1200  # Максимальная комфортная ширина
        if self.scrollable_matrix_frame.winfo_reqwidth() > max_matrix_width:
            for col in range(len(heights) + 1):
                self.scrollable_matrix_frame.columnconfigure(col, minsize=80)  # Фиксируем ширину столбцов

        # Обновляем размеры окна после изменения матрицы
        self.adjust_window_size()

    def create_cell(self, sheet_name, data, length, height, row, col):
        pattern = f"/{height}/{length}"
        match = data[data['Наименование'].str.contains(pattern, na=False)]

        if not match.empty:
            product = match.iloc[0]
            art = str(product['Артикул']).strip()

            value = self.entry_values.get((sheet_name, art), "")

            entry = tk.Entry(
                self.scrollable_matrix_frame,
                width=8,
                justify="center",
                bg='#e6f3ff' if self.has_any_value() else 'white',
                relief='solid',
                borderwidth=1,
                validate='key',
                validatecommand=(self.root.register(self.validate_input), '%P')
            )

            entry.insert(0, value)
            
            entry.bind("<FocusIn>", lambda e: self.on_entry_focus_in(e))
            entry.bind("<FocusOut>", lambda e, s=sheet_name, a=art: self.on_entry_focus_out(e, s, a))
            entry.bind("<Return>", lambda e, s=sheet_name, a=art: self.on_entry_focus_out(e, s, a))
            entry.bind("<Tab>", lambda e, s=sheet_name, a=art: self.on_entry_focus_out(e, s, a))

            entry.bind("<Enter>", lambda e, p=product: self.show_tooltip_on_hover(p))
            entry.bind("<Leave>", lambda e: self.hide_tooltip_on_leave())

            self.entries[(sheet_name, art)] = entry
            entry.grid(row=row, column=col, sticky="nsew", padx=1, pady=1)

    def on_entry_focus_in(self, event):
        """Обработчик получения фокуса Entry"""
        entry = event.widget
        text = entry.get()
        
        # Если текст содержит '+' (длинная формула) или если текст длиннее 3 символов
        if '+' in text or len(text) > 3:
            # Выравниваем текст по правому краю
            entry.config(justify="right")
            # Устанавливаем курсор в конец текста
            entry.icursor(len(text))
            # Прокручиваем так, чтобы текст был виден полностью
            entry.xview_moveto(1.0)
        else:
            # Для коротких значений или пустых ячеек - центрируем
            entry.config(justify="center")
            if text:  # Если есть текст, курсор в конце
                entry.icursor(len(text))

    def set_cursor_and_scroll(self, entry, text):
        # Устанавливаем курсор в конец текста
        entry.icursor(len(text))
        
        # Прокручиваем так, чтобы последний символ был виден
        entry.xview_moveto(1.0)
        
        # Делаем курсор видимым
        entry.focus_set()

    def on_entry_focus_out(self, event, sheet_name, art):
        """Обработчик потери фокуса Entry"""
        entry = event.widget
        text = entry.get()
        
        # Всегда возвращаем центрирование текста при потере фокуса
        entry.config(justify="center")
        
        # Прокручиваем текст так, чтобы были видны последние символы
        if text:
            entry.xview_moveto(1.0)
        
        # Сохраняем значение
        if text:
            self.entry_values[(sheet_name, art)] = text
        else:
            self.entry_values.pop((sheet_name, art), None)
        
        # Обновляем цвета всех ячеек
        color = '#e6f3ff' if self.has_any_value() else 'white'
        for e in list(self.entries.values()):  # Создаем копию списка
            try:
                if e.winfo_exists():  # Проверяем, существует ли виджет
                    e.config(bg=color)
            except tk.TclError:
                continue

    def show_tooltip_on_hover(self, product):
        """Показывает подсказку при наведении, если включен чекбокс"""
        if not self.show_tooltips_var.get():
            return
            
        # Создаем подсказку, если ее еще нет
        if not hasattr(self, '_hover_tooltip'):
            self._hover_tooltip = tk.Toplevel(self.root)
            self._hover_tooltip.wm_overrideredirect(True)
            self._hover_tooltip.withdraw()
            
            self._hover_tooltip_label = ttk.Label(
                self._hover_tooltip,
                background="#ffffe0",
                relief="solid",
                padding=5,
                font=("Segoe UI", 9)
            )
            self._hover_tooltip_label.pack()
        
        # Формируем текст подсказки
        power = product.get('Мощность, Вт', '')
        power_text = f"Мощность: {power} Вт" if power else "Мощность: не указана"
        
        text = (f"Артикул: {product['Артикул']}\n"
                f"{power_text}\n"
                f"Вес: {product['Вес, кг']} кг\n"
                f"Объем: {product['Объем, м3']} м³")
        
        self._hover_tooltip_label.config(text=text)
        
        # Позиционируем подсказку рядом с курсором
        x = self.root.winfo_pointerx() + 15
        y = self.root.winfo_pointery() + 15
        self._hover_tooltip.wm_geometry(f"+{x}+{y}")
        self._hover_tooltip.deiconify()   

    def hide_tooltip_on_leave(self):
        """Скрывает подсказку при уходе курсора"""
        if hasattr(self, '_hover_tooltip'):
            self._hover_tooltip.withdraw()             

    def save_value(self, sheet_name, art):
        """Сохраняет значение из поля ввода"""
        entry = self.entries.get((sheet_name, art))
        if entry:
            value = entry.get()
            if value:
                self.entry_values[(sheet_name, art)] = value
            else:
                self.entry_values.pop((sheet_name, art), None)

    def create_tooltip(self, widget, product):
        # Создаем подсказку и сохраняем ссылку на виджете
        tooltip = tk.Toplevel(self.root)
        tooltip.wm_overrideredirect(True)
        tooltip.withdraw()
        widget._tooltip = tooltip  # Сохраняем ссылку на подсказку
        
        power = product.get('Мощность, Вт', '')
        power_text = f"Мощность: {power} Вт" if power else "Мощность: не указана"
        
        text = (f"Артикул: {product['Артикул']}\n"
                f"{power_text}\n"
                f"Вес: {product['Вес, кг']} кг\n"
                f"Объем: {product['Объем, м3']} м³")
        
        label = ttk.Label(tooltip, text=text, background="#ffffe0", relief="solid", padding=5)
        label.pack()
        
        widget.bind("<Enter>", lambda e: self.show_tooltip(tooltip, widget))
        widget.bind("<Leave>", lambda e: tooltip.withdraw())

    def show_tooltip(self, tooltip, widget):
        x = widget.winfo_rootx() + 25
        y = widget.winfo_rooty() + 25
        tooltip.wm_geometry(f"+{x}+{y}")
        tooltip.deiconify()

    def calculate_brackets(self, radiator_type, length, height, bracket_type, qty_radiator=1):
        """
        Рассчитывает необходимые кронштейны для радиатора
        
        Параметры:
            radiator_type (str): Тип радиатора ("10", "11", "20" и т.д.)
            length (int): Длина радиатора в мм (400-2000)
            height (int): Высота радиатора в мм (300,400,500,600,900)
            bracket_type (str): Тип крепления
            qty_radiator (int): Количество радиаторов (по умолчанию 1)
        
        Возвращает:
            list: Список кортежей (артикул, количество)
        """
        brackets = []
        
        # Настенные кронштейны
        if bracket_type == "Настенные кронштейны":
            if radiator_type in ["10", "11"]:
                brackets.extend([
                    ("К9.2L", 2 * qty_radiator),
                    ("К9.2R", 2 * qty_radiator)
                ])
                if 1700 <= length <= 2000:
                    brackets.append(("К9.3-40", 1 * qty_radiator))
            
            elif radiator_type in ["20", "21", "22", "30", "33"]:
                art_map = {
                    300: "К15.4300",
                    400: "К15.4400", 
                    500: "К15.4500",
                    600: "К15.4600",
                    900: "К15.4900"
                }
                if height in art_map:
                    art = art_map[height]
                    if 400 <= length <= 1600:
                        qty = 2 * qty_radiator
                    elif 1700 <= length <= 2000:
                        qty = 3 * qty_radiator
                    else:
                        qty = 0
                    if qty > 0:
                        brackets.append((art, qty))
        
        # Напольные кронштейны
        elif bracket_type == "Напольные кронштейны":
            if radiator_type in ["10", "11"]:
                if 300 <= height <= 400:
                    main_art = "КНС450"
                elif 500 <= height <= 600:
                    main_art = "КНС470" 
                elif height == 900:
                    main_art = "КНС4100"
                else:
                    main_art = None
                
                if main_art:
                    brackets.append((main_art, 2 * qty_radiator))
                    if 1700 <= length <= 2000:
                        brackets.append(("КНС430", 1 * qty_radiator))
            
            elif radiator_type == "21":
                if 300 <= height <= 400:
                    art = "КНС650"
                elif 500 <= height <= 600:
                    art = "КНС670"
                elif height == 900:
                    art = "КНС6100"
                else:
                    art = None
                
                if art:
                    if 400 <= length <= 1000:
                        qty = 2 * qty_radiator
                    elif 1100 <= length <= 1600:
                        qty = 3 * qty_radiator
                    elif 1700 <= length <= 2000:
                        qty = 4 * qty_radiator
                    else:
                        qty = 0
                    if qty > 0:
                        brackets.append((art, qty))
            
            elif radiator_type in ["20", "22", "30", "33"]:
                if 300 <= height <= 400:
                    art = "КНС550"
                elif 500 <= height <= 600:
                    art = "КНС570"
                elif height == 900:
                    art = "КНС5100"
                else:
                    art = None
                
                if art:
                    if 400 <= length <= 1000:
                        qty = 2 * qty_radiator
                    elif 1100 <= length <= 1600:
                        qty = 3 * qty_radiator
                    elif 1700 <= length <= 2000:
                        qty = 4 * qty_radiator
                    else:
                        qty = 0
                    if qty > 0:
                        brackets.append((art, qty))
        
        return brackets

    def create_context_menu(self, tree, spec_data):
        """Создает контекстное меню для удаления строк"""
        context_menu = tk.Menu(tree, tearoff=0)
        context_menu.add_command(
            label="Удалить",
            command=lambda: self.delete_selected_row(tree, spec_data)
        )
        context_menu.add_separator()  # Добавляем разделитель

        def show_context_menu(event):
            item = tree.identify_row(event.y)
            if item:
                tree.selection_set(item)
                context_menu.post(event.x_root, event.y_root)
        
        tree.bind("<Button-3>", show_context_menu)

    def update_treeview(self, tree, spec_data):
        """Обновляет данные в Treeview"""
        tree.delete(*tree.get_children())
        for _, row in spec_data.iterrows():
            formatted_row = [
                row["№"],
                row["Артикул"],
                row["Наименование"],
                row["Мощность, Вт"],
                f"{float(row['Цена, руб (с НДС)']):.2f}",
                f"{float(row['Скидка, %']):.2f}",
                f"{float(row['Цена со скидкой, руб (с НДС)']):.2f}",
                row["Кол-во"],
                f"{float(row['Сумма, руб (с НДС)']):.2f}"
            ]
            tree.insert("", "end", values=formatted_row)
        
        # Добавляем обновленную итоговую строку
        total_sum = spec_data["Сумма, руб (с НДС)"].sum()
        total_qty_radiators = sum(spec_data.query("Наименование.str.contains('Радиатор')")["Кол-во"].apply(self.parse_quantity))
        total_qty_brackets = sum(spec_data.query("Наименование.str.contains('Кронштейн')")["Кол-во"].apply(self.parse_quantity))
        
        tree.insert("", "end", values=[
            "Итого", "", "", "", "", "", "",
            f"{total_qty_radiators} / {total_qty_brackets}",
            f"{total_sum:.2f}"
        ], tags=("total",))
        tree.tag_configure("total", background="#e0e0e0", font=("Segoe UI", 9, "bold")) 

    def delete_selected_row(self, tree, spec_data):
        """Удаляет выбранную строку из Treeview и данных"""
        selected_item = tree.selection()
        if selected_item:
            item_values = tree.item(selected_item)['values']
            if item_values and item_values[0] != "Итого":  # Нельзя удалить итоговую строку
                # Удаляем из DataFrame
                index_to_remove = item_values[0] - 1  # № начинается с 1
                if 0 <= index_to_remove < len(spec_data):
                    spec_data.drop(index_to_remove, inplace=True)
                    spec_data.reset_index(drop=True, inplace=True)
                    # Обновляем номера строк
                    spec_data["№"] = range(1, len(spec_data) + 1)
                    
                    # Обновляем Treeview
                    self.update_treeview(tree, spec_data)        

    def prepare_spec_data(self):
        # Сохраняем значение из текущей активной ячейки (если есть)
        if self.root.focus_get() in self.entries.values():
            for (sheet_name, art), entry in self.entries.items():
                if entry == self.root.focus_get():
                    value = entry.get()
                    if value:
                        self.entry_values[(sheet_name, art)] = value
                    else:
                        self.entry_values.pop((sheet_name, art), None)
                    break
        spec_data = []
        radiator_data = []
        bracket_data = []
        brackets_temp = {}

        # Обработка радиаторов
        for (sheet_name, art), value in self.entry_values.items():
            if value and sheet_name in self.sheets:
                try:
                    # Получаем значение из entry_values (может быть "1+3")
                    raw_value = self.entry_values.get((sheet_name, art), "")
                    # Вычисляем сумму только при формировании спецификации
                    qty_radiator = self.parse_quantity(raw_value)
                    mask = self.sheets[sheet_name]['Артикул'] == art
                    product = self.sheets[sheet_name].loc[mask]
                    
                    if product.empty:
                        continue
                    
                    product = product.iloc[0]
                    radiator_type = sheet_name.split()[-1]
                    price = float(product['Цена, руб'])
                    # Получаем скидку из переменной интерфейса
                    discount = float(self.radiator_discount_var.get()) if self.radiator_discount_var.get() else 0.0
                    discounted_price = round(price * (1 - discount / 100), 2)
                    total = round(discounted_price * qty_radiator, 2)
                    
                    # Извлекаем параметры для сортировки из наименования
                    name_parts = product['Наименование'].split('/')
                    height = int(name_parts[-2].replace('мм', '').strip())
                    length = int(name_parts[-1].replace('мм', '').strip().split()[0])
                    
                    # Определяем Вид подключения для сортировки
                    connection_type = "VK" if "VK" in sheet_name else "K"
                    
                    radiator_data.append({
                        "№": len(radiator_data) + 1,
                        "Артикул": str(product['Артикул']).strip(),
                        "Наименование": str(product['Наименование']),
                        "Мощность, Вт": float(product.get('Мощность, Вт', 0)),
                        "Цена, руб (с НДС)": float(price),
                        "Скидка, %": float(discount),
                        "Цена со скидкой, руб (с НДС)": float(discounted_price),
                        "Кол-во": int(qty_radiator),
                        "Сумма, руб (с НДС)": float(total),
                        "ConnectionType": connection_type,  # Для группировки VK/K
                        "RadiatorType": int(radiator_type),  # Тип радиатора (10, 11, 20 и т.д.)
                        "Height": height,  # Высота для сортировки
                        "Length": length  # Длина для сортировки
                    })

                    # Обработка кронштейнов (только если не добавлены в предпросмотре)
                    if self.bracket_var.get() != "Без кронштейнов" and not hasattr(self, 'preview_brackets_added'):
                        brackets = self.calculate_brackets(
                            radiator_type=radiator_type,
                            length=length,
                            height=height,
                            bracket_type=self.bracket_var.get(),
                            qty_radiator=qty_radiator
                        )
                        
                        for art_bracket, qty_bracket in brackets:
                            mask_bracket = self.brackets_df['Артикул'] == art_bracket
                            bracket_info = self.brackets_df.loc[mask_bracket]
                            
                            if bracket_info.empty:
                                continue
                                
                            key = art_bracket.strip()
                            if key not in brackets_temp:
                                brackets_temp[key] = {
                                    "Артикул": art_bracket,
                                    "Наименование": str(bracket_info.iloc[0]['Наименование']),
                                    "Цена, руб (с НДС)": float(bracket_info.iloc[0]['Цена, руб']),
                                    "Кол-во": 0,
                                    "Сумма, руб (с НДС)": 0.0
                                }
                            
                            price_bracket = float(bracket_info.iloc[0]['Цена, руб'])
                            # Получаем скидку на кронштейны из переменной интерфейса
                            discount_bracket = float(self.bracket_discount_var.get()) if self.bracket_discount_var.get() else 0.0
                            discounted_price_bracket = round(price_bracket * (1 - discount_bracket / 100), 2)
                            qty_total = qty_bracket
                            
                            brackets_temp[key]["Кол-во"] += int(qty_total)
                            brackets_temp[key]["Сумма, руб (с НДС)"] += round(discounted_price_bracket * qty_total, 2)

                except Exception as e:
                    messagebox.showerror("Ошибка", f"Ошибка в данных радиатора: {str(e)}")
                    return None

        # Формирование данных кронштейнов
        if brackets_temp:
            for b in brackets_temp.values():
                bracket_discount = float(self.bracket_discount_var.get()) if self.bracket_discount_var.get() else 0.0
                price_with_discount = round(float(b["Цена, руб (с НДС)"]) * (1 - bracket_discount / 100), 2)
                
                bracket_data.append({
                    "№": len(radiator_data) + len(bracket_data) + 1,
                    "Артикул": str(b["Артикул"]),
                    "Наименование": str(b["Наименование"]),
                    "Мощность, Вт": 0.0,
                    "Цена, руб (с НДС)": float(b["Цена, руб (с НДС)"]),
                    "Скидка, %": float(bracket_discount),
                    "Цена со скидкой, руб (с НДС)": float(price_with_discount),
                    "Кол-во": int(b["Кол-во"]),
                    "Сумма, руб (с НДС)": float(b["Сумма, руб (с НДС)"]),
                    "ConnectionType": "Bracket"  # Для кронштейнов
                })

        radiator_data_sorted = sorted(
            radiator_data, 
            key=lambda x: (
                0 if x["ConnectionType"] == "VK" else 1,  # Сначала VK, потом K
                x["RadiatorType"],  # Затем по типу радиатора (10, 11, 20...)
                x["Height"],  # Затем по высоте
                x["Length"]  # Затем по длине
            )
        )
        
        # Обновляем номера после сортировки
        for i, item in enumerate(radiator_data_sorted, 1):
            item["№"] = i
        
        # Объединение данных (отсортированные радиаторы + кронштейны)
        combined_data = radiator_data_sorted + bracket_data
        
        if not combined_data:
            messagebox.showwarning("Пусто", "Нет данных для формирования спецификации")
            return None

        # Создание DataFrame (удаляем временные поля для сортировки)
        df = pd.DataFrame(
            combined_data,
            columns=[
                "№", "Артикул", "Наименование", "Мощность, Вт",
                "Цена, руб (с НДС)", "Скидка, %",
                "Цена со скидкой, руб (с НДС)", "Кол-во",
                "Сумма, руб (с НДС)"
            ]
        )
        
        return df

    def load_excel_spec(self):
        """Загружает данные из Excel-спецификации, автоматически находя нужные столбцы"""
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if not file_path:
            return

        try:
            # Определяем движок для чтения
            if file_path.endswith('.xlsx'):
                engine = 'openpyxl'
            elif file_path.endswith('.xls'):
                engine = 'xlrd'
            else:
                messagebox.showerror("Ошибка", "Неподдерживаемый формат файла")
                return

            # Читаем весь файл для анализа
            df = pd.read_excel(file_path, engine=engine, header=None)

            # 1. Находим столбцы с артикулами и количеством
            art_col = None
            qty_col = None
            
            # Перебираем все строки сверху вниз
            for i, row in df.iterrows():
                # Ищем в текущей строке нужные заголовки
                for j, cell in enumerate(row):
                    cell_str = str(cell).strip().lower()
                    
                    # Проверяем возможные варианты названий столбцов
                    if not art_col and any(x in cell_str for x in ['артикул', 'art', 'код']):
                        art_col = j
                    if not qty_col and any(x in cell_str for x in ['кол-во', 'количество', 'qty']):
                        qty_col = j
                
                # Если нашли оба столбца - выходим из цикла
                if art_col is not None and qty_col is not None:
                    header_row = i  # Запоминаем строку с заголовками
                    break
            
            # Если не нашли нужные столбцы - берем первые два столбца
            if art_col is None:
                art_col = 0
            if qty_col is None:
                qty_col = 1 if len(df.columns) > 1 else 0

            # 2. Читаем данные, начиная со строки после найденных заголовков
            data_rows = []
            for i in range(header_row + 1, len(df)):
                art = str(df.iloc[i, art_col]).strip()
                qty = df.iloc[i, qty_col]
                
                # Пропускаем пустые строки и строки "Итого"
                if not art or art.lower() == 'итого':
                    continue
                    
                try:
                    qty = float(qty)
                    if qty > 0:
                        data_rows.append((art, int(qty)))
                except (ValueError, TypeError):
                    continue

            # 3. Очищаем текущие значения и загружаем новые
            self.entry_values.clear()
            total_loaded = 0
            total_qty = 0
            
            for art, qty in data_rows:
                # Ищем артикул в данных программы
                found = False
                for sheet_name, sheet_data in self.sheets.items():
                    if art in sheet_data['Артикул'].astype(str).str.strip().values:
                        self.entry_values[(sheet_name, art)] = str(qty)
                        total_loaded += 1
                        total_qty += qty
                        found = True
                        break
                
                if not found:
                    print(f"Артикул не найден: {art}")

            # 4. Обновляем интерфейс
            self.refresh_matrix()
            self.global_highlight()
            
            messagebox.showinfo(
                "Успех",
            f"   Обработано строк: {len(df)}\n"
            f"   Загружено артикулов: {total_loaded}\n"
            f"   Общее количество радиаторов: {total_qty}"
            )

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить файл:\n{str(e)}")
            print(f"Ошибка загрузки: {traceback.format_exc()}")

    def load_csv_spec(self):
        file_path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
        if not file_path:
            return

        try:
            encodings = ['utf-8-sig', 'cp1251', 'windows-1251', 'iso-8859-1']
            df = None
            
            for encoding in encodings:
                try:
                    # Читаем файл без заголовков сначала
                    df = pd.read_csv(file_path, sep=';', encoding=encoding, header=None)
                    
                    # Проверяем, есть ли заголовки
                    has_headers = not df.iloc[0, 0].replace('.', '').isdigit()
                    
                    if has_headers:
                        # Читаем снова с заголовками
                        df = pd.read_csv(file_path, sep=';', encoding=encoding, header=0)
                    else:
                        # Берем первые два столбца как артикул и количество
                        df.columns = ['Артикул', 'Кол-во']
                    
                    df = df[df.iloc[:, 0].notna()]
                    break
                except UnicodeDecodeError:
                    continue
            
            if df is None:
                raise ValueError("Не удалось определить кодировку файла")

            # Определяем индексы столбцов с артикулами и количеством
            art_col = None
            qty_col = None
            
            # Ищем столбцы по возможным названиям
            for col in df.columns:
                col_lower = str(col).lower()
                if 'артикул' in col_lower or 'art' in col_lower or 'код' in col_lower:
                    art_col = col
                elif 'кол-во' in col_lower or 'количество' in col_lower or 'qty' in col_lower:
                    qty_col = col
            
            # Если не нашли стандартные названия - берем первые два столбца
            if art_col is None:
                art_col = df.columns[0]
            if qty_col is None:
                qty_col = df.columns[1] if len(df.columns) > 1 else None

            if qty_col is None:
                messagebox.showerror("Ошибка", "Не найден столбец с количеством")
                return

            # Удаляем пробелы из артикулов
            df[art_col] = df[art_col].astype(str).str.replace(' ', '').str.strip()
            
            # Преобразуем количество в целые числа
            df[qty_col] = df[qty_col].astype(str).str.replace('.0', '').str.strip()
            df[qty_col] = pd.to_numeric(df[qty_col], errors='coerce').fillna(0).astype(int)
            
            # Очищаем текущие значения
            self.entry_values.clear()
            
            total_qty_radiators = 0
            total_qty_brackets = 0

            # Группируем по артикулам и суммируем количества
            grouped_df = df.groupby(art_col)[qty_col].sum().reset_index()

            for _, row in grouped_df.iterrows():
                art = str(row[art_col]).strip()
                qty = int(row[qty_col])  # Гарантированно целое число

                # Ищем артикул во всех листах матрицы
                found = False
                for sheet_name, sheet_data in self.sheets.items():
                    mask = sheet_data['Артикул'].astype(str).str.strip() == art
                    if mask.any():
                        product = sheet_data[mask].iloc[0]
                        # Сохраняем как строку без .0
                        self.entry_values[(sheet_name, art)] = str(int(qty))
                        total_qty_radiators += qty
                        found = True
                        break

                if not found:
                    print(f"Артикул не найден в матрице: {art}")

            # Полностью пересоздаем матрицу с новыми значениями
            self.show_selected_matrix()
            self.global_highlight()

            messagebox.showinfo("Успех", f"""
            Загружено строк: {len(df)}
            Всего радиаторов: {total_qty_radiators}
            """)

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка загрузки CSV: {str(e)}")

    def highlight_filled_cells(self):
        """Подсвечивает заполненные ячейки"""
        has_values = False
        
        # Проверяем, есть ли хотя бы одно заполненное поле
        for (sheet_name, art), entry in self.entries.items():
            value = self.entry_values.get((sheet_name, art), "")
            if value:
                has_values = True
                break
        
        # Окрашиваем все ячейки
        for (sheet_name, art), entry in self.entries.items():
            value = self.entry_values.get((sheet_name, art), "")
            if value:
                # Заполненные ячейки - желтый цвет
                entry.config(background='#e6f3ff')  # Желтый цвет для заполненных ячеек
            else:
                if has_values:
                    # Незаполненные ячейки при наличии заполненных - голубой цвет
                    entry.config(background='#e6f3ff')
                else:
                    # Если нет заполненных ячеек - белый фон
                    entry.config(background='white')

    def reset_cell_colors(self):
        """Возвращает стандартный цвет всем ячейкам"""
        for entry in self.entries.values():
            if hasattr(entry, 'original_bg'):
                entry.config(background=entry.original_bg)
            else:
                entry.config(background='white')

    def update_treeview(self, tree, spec_data):
        """Обновляет данные в Treeview"""
        tree.delete(*tree.get_children())
        for _, row in spec_data.iterrows():
            formatted_row = [
                row["№"],
                row["Артикул"],
                row["Наименование"],
                row["Мощность, Вт"],
                f"{float(row['Цена, руб (с НДС)']):.2f}",
                f"{float(row['Скидка, %']):.2f}",
                f"{float(row['Цена со скидкой, руб (с НДС)']):.2f}",
                row["Кол-во"],
                f"{float(row['Сумма, руб (с НДС)']):.2f}"
            ]
            tree.insert("", "end", values=formatted_row)
        
        # Добавляем итоговую строку
        total_sum = spec_data["Сумма, руб (с НДС)"].sum()
        total_qty_radiators = sum(spec_data.query("Наименование.str.contains('Радиатор')")["Кол-во"].apply(self.parse_quantity))
        total_qty_brackets = sum(spec_data.query("Наименование.str.contains('Кронштейн')")["Кол-во"].apply(self.parse_quantity))
        
        tree.insert("", "end", values=[
            "Итого", "", "", "", "", "", "",
            f"{total_qty_radiators} / {total_qty_brackets}",
            f"{total_sum:.2f}"
        ], tags=("total",))
        tree.tag_configure("total", background="#e0e0e0", font=("Segoe UI", 9, "bold"))             

    def generate_spec(self, file_type="excel", tree=None):
        # Если вызывается из окна предпросмотра, используем сохраненные данные
        if hasattr(self, '_current_spec_data') and self._current_spec_data is not None:
            spec_data = self._current_spec_data
        else:
            # Иначе готовим данные как обычно
            spec_data = self.prepare_spec_data()
        
        if spec_data is None or spec_data.empty:
            messagebox.showwarning("Пусто", "Нет данных для сохранения")
            return

        if file_type == "excel":
            try:
                file_name = "Расчёт стоимости.xlsx"
                temp_dir = tempfile.gettempdir()
                
                # Генерируем уникальное имя файла, если файл уже существует
                counter = 1
                base_name = "Расчёт стоимости"
                file_path = os.path.join(temp_dir, f"{base_name}.xlsx")
                
                while os.path.exists(file_path):
                    file_path = os.path.join(temp_dir, f"{base_name}_{counter}.xlsx")
                    counter += 1
                    if counter > 10000000000000000000000000:  # Защита от бесконечного цикла
                        raise Exception("Не удалось создать уникальное имя файла")
                
                # Инициализируем correspondence_data как None по умолчанию
                correspondence_data = None
                # Проверяем, есть ли данные о переподборе
                if hasattr(self, '_correspondence_df') and self._correspondence_df is not None:
                    correspondence_data = self._correspondence_df
                
                self.save_excel_spec(spec_data, file_path, correspondence_data)
                
                # Попытка открыть файл с обработкой возможных ошибок
                try:
                    self.open_file_default_app(file_path)
                except Exception as open_error:
                    # Если не удалось открыть, предлагаем пользователю открыть вручную
                    if messagebox.askyesno(
                        "Ошибка открытия",
                        f"Не удалось автоматически открыть файл:\n{str(open_error)}\n\n"
                        f"Файл сохранен по пути:\n{file_path}\n\n"
                        "Хотите открыть его вручную?"
                    ):
                        try:
                            os.startfile(os.path.dirname(file_path))
                        except:
                            messagebox.showinfo(
                                "Информация",
                                f"Файл сохранен по пути:\n{file_path}\n\n"
                                "Откройте его вручную."
                            )
                
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при создании Excel:\n{str(e)}")
                
        elif file_type == "csv":
            file_path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV Files", "*.csv")],
                title="Сохранить как CSV"
            )
            if not file_path:
                return

            try:
                df = pd.DataFrame({
                    "Артикул": spec_data["Артикул"],
                    "Кол-во": spec_data["Кол-во"]
                })
                df.to_csv(file_path, index=False, sep=';', encoding='utf-8-sig')
                messagebox.showinfo("Успешно", f"CSV сохранен:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка сохранения CSV:\n{str(e)}")

    def save_excel_spec(self, spec_data, path, correspondence_data=None):
        """Сохраняет спецификацию в Excel с сортировкой по типоразмеру"""
        from openpyxl.styles import Font, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter

        # Создаем книгу и лист
        wb = Workbook()
        ws = wb.active
        ws.title = "Спецификация"

        # Стили оформления
        header_font = Font(name='Calibri', size=11, bold=True)
        data_font = Font(name='Calibri', size=11)
        bold_font = Font(name='Calibri', size=11, bold=True)
        alignment_center = Alignment(horizontal='center', vertical='center')
        alignment_left = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'),
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Стиль для денежных значений (2 знака после запятой)
        money_style = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        # Заголовки столбцов
        headers = [
            "№", "Артикул", "Наименование", "Мощность, Вт",
            "Цена, руб (с НДС)", "Скидка, %", 
            "Цена со скидкой, руб (с НДС)", "Кол-во", 
            "Сумма, руб (с НДС)"
        ]
        ws.append(headers)

        # Применяем стили к заголовкам
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = alignment_center
            cell.border = thin_border

        # Разделяем радиаторы и кронштейны
        radiator_data = []
        bracket_data = []
        
        for _, row in spec_data.iterrows():
            if "Кронштейн" in str(row["Наименование"]):
                bracket_data.append(row)
            else:
                radiator_data.append(row)
        
        # Функция для извлечения параметров сортировки из наименования
        def get_sort_key(row):
            name = str(row["Наименование"])
            # Определяем тип подключения
            connection_type = 0 if "VK" in name else 1  # Сначала VK, потом K
            
            # Извлекаем тип радиатора (10, 11, 20 и т.д.)
            radiator_type = 0
            if "тип 10" in name: radiator_type = 10
            elif "тип 11" in name: radiator_type = 11
            elif "тип 20" in name: radiator_type = 20
            elif "тип 21" in name: radiator_type = 21
            elif "тип 22" in name: radiator_type = 22
            elif "тип 30" in name: radiator_type = 30
            elif "тип 33" in name: radiator_type = 33
            
            # Извлекаем высоту и длину
            parts = name.split('/')
            height = int(parts[-2].replace('мм', '').strip())
            length = int(parts[-1].replace('мм', '').strip().split()[0])
            
            return (connection_type, radiator_type, height, length)
        
        # Сортировка радиаторов
        radiator_data_sorted = sorted(radiator_data, key=get_sort_key)
        
        # Объединение данных (отсортированные радиаторы + кронштейны)
        combined_data = radiator_data_sorted + bracket_data
        
        # Заполняем Excel
        for i, row in enumerate(combined_data, 2):  # Начинаем с 2 строки
            # Для кронштейнов заменяем 0 на пустую строку в столбце мощности
            power_value = "" if "Кронштейн" in str(row["Наименование"]) else row["Мощность, Вт"]
            
            ws.append([
                i-1,  # №
                str(row["Артикул"]),  # Артикул как строка
                row["Наименование"],
                power_value,
                float(row['Цена, руб (с НДС)']),
                float(row['Скидка, %']),
                float(row['Цена со скидкой, руб (с НДС)']),
                int(row['Кол-во']),
                float(row['Сумма, руб (с НДС)'])
            ])

            # Форматируем строки данных
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=i, column=col)
                cell.font = data_font
                cell.border = thin_border
                
                # Устанавливаем числовые форматы
                if col in [5, 7, 9]:  # Столбцы с ценами и суммами
                    cell.number_format = money_style
                    cell.alignment = alignment_center
                elif col == 4:  # Столбец "Мощность, Вт" - центрируем
                    cell.alignment = alignment_center
                elif col in [1, 6, 8]:  # Другие числовые столбцы
                    cell.alignment = alignment_center
                else:
                    cell.alignment = alignment_left

        # Добавляем итоговую строку
        total_row = len(combined_data) + 2
        total_sum = spec_data["Сумма, руб (с НДС)"].sum()
        total_qty_radiators = sum(spec_data.query("Наименование.str.contains('Радиатор')")["Кол-во"].apply(self.parse_quantity))
        total_qty_brackets = sum(spec_data.query("Наименование.str.contains('Кронштейн')")["Кол-во"].apply(self.parse_quantity))
        
        ws.append(["Итого", "", "", "", "", "", "", f"{total_qty_radiators}/{total_qty_brackets}", total_sum])
        
        # Форматируем итоговую строку
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=total_row, column=col)
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = alignment_center
            if col in [5, 7, 9]:  # Форматируем денежные столбцы
                cell.number_format = money_style

        # Добавляем вес и объем
        total_weight, total_volume = self.calculate_total_weight_and_volume(spec_data)
        
        # Пустая строка
        ws.append([])

        # Строка с весом
        ws.append([f"Суммарный вес радиаторов без учета упаковки и кронштейнов- {total_weight} кг."])
        ws.merge_cells(start_row=total_row + 2, start_column=1, end_row=total_row + 2, end_column=9)
        cell = ws.cell(row=total_row + 2, column=1)
        cell.font = Font(name='Calibri', size=11)
        cell.alignment = alignment_left

        # Строка с объемом
        ws.append([f"Суммарный объем радиаторов без учета упаковки и кронштейнов- {total_volume} м3."])
        ws.merge_cells(start_row=total_row + 3, start_column=1, end_row=total_row + 3, end_column=9)
        cell = ws.cell(row=total_row + 3, column=1)
        cell.font = Font(name='Calibri', size=11)
        cell.alignment = alignment_left

        # Настраиваем ширину столбцов для основного листа
        column_widths = {
            'A': 5,    # №
            'B': 12,   # Артикул
            'C': 60,   # Наименование
            'D': 15,   # Мощность
            'E': 20,   # Цена
            'F': 10,   # Скидка
            'G': 30,   # Цена со скидкой
            'H': 10,   # Кол-во
            'I': 20    # Сумма
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Добавляем лист с таблицей соответствия, если есть данные
        if correspondence_data is not None and not correspondence_data.empty:
            ws_corr = wb.create_sheet("Таблица соответствия")
            
            # Заголовки таблицы соответствия
            corr_headers = list(correspondence_data.columns)
            ws_corr.append(corr_headers)
            
            # Применяем стили к заголовкам
            for col in range(1, len(corr_headers) + 1):
                cell = ws_corr.cell(row=1, column=col)
                cell.font = header_font
                cell.alignment = alignment_center
                cell.border = thin_border
            
            # Заполняем данные
            for i, row in correspondence_data.iterrows():
                ws_corr.append(list(row))
                
                # Форматируем строки данных
                for col in range(1, len(corr_headers) + 1):
                    cell = ws_corr.cell(row=i+2, column=col)
                    cell.font = data_font
                    cell.border = thin_border
                    # Центрируем только столбец "Количество" (2-й столбец)
                    if col == 2:
                        cell.alignment = alignment_center
                    else:
                        cell.alignment = alignment_left
            
            # Автоподбор ширины столбцов для листа соответствия
            for col_idx, column_name in enumerate(corr_headers, 1):
                max_length = len(str(column_name))  # Начинаем с длины заголовка
                column_letter = get_column_letter(col_idx)
                
                # Ищем максимальную длину содержимого в столбце
                for row in ws_corr.iter_rows(min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                
                # Устанавливаем ширину с небольшим запасом
                adjusted_width = (max_length + 2) * 1.2
                ws_corr.column_dimensions[column_letter].width = adjusted_width

        # Сохраняем файл
        wb.save(path)

    def calculate_totals(self, spec_data):
        total_weight = 0.0
        total_volume = 0.0
        
        for _, row in spec_data.iterrows():
            art = str(row['Артикул']).strip()
            qty = int(row['Кол-во'])
            
            for sheet, data in self.sheets.items():
                data_arts = data['Артикул'].astype(str).str.strip()
                product = data[data_arts == art]
                if not product.empty:
                    total_weight += float(product.iloc[0]['Вес, кг']) * qty
                    total_volume += float(product.iloc[0]['Объем, м3']) * qty
                    break
        
        return float(total_weight), float(total_volume)
    
    def create_header_tooltip(self, tree, column, text):
        """Создает подсказку для конкретного заголовка столбца"""
        def show_tooltip(event):
            # Получаем координаты заголовка
            x = tree.winfo_rootx() + tree.column(column, 'x') 
            y = tree.winfo_rooty() - 20  # Сдвигаем выше заголовка
            
            # Создаем подсказку
            tooltip = tk.Toplevel(tree)
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{x}+{y}")
            
            label = ttk.Label(
                tooltip, 
                text=text, 
                background="#ffffe0", 
                relief="solid", 
                padding=5,
                font=("Segoe UI", 9)
            )
            label.pack()
            
            tooltip.after(300, tooltip.destroy)  # Автоматическое закрытие через 0.3 сек

        # Привязываем события
        tree.tag_bind(column, "<Enter>", show_tooltip)

    
    def create_header_tooltip(self, widget, text):
        """Создает подсказку для заголовка столбца"""
        tooltip = tk.Toplevel(self.root)
        tooltip.wm_overrideredirect(True)
        tooltip.withdraw()
        
        label = ttk.Label(
            tooltip, 
            text=text, 
            background="#ffffe0", 
            relief="solid", 
            padding=5,
            font=("Segoe UI", 9)
        )
        label.pack()
        
        widget.bind("<Enter>", lambda e: self.show_header_tooltip(tooltip, widget))
        widget.bind("<Leave>", lambda e: tooltip.withdraw())

    def format_power(self, power_w):
        """
        Форматирует мощность с автоматическим выбором единиц измерения:
        - Если больше или равно 1 000 000 Вт -> переводит в МВт
        - Если больше или равно 1 000 Вт -> переводит в кВт
        - Меньше 1 000 Вт -> оставляет в Вт
        Возвращает строку с единицей измерения.
        """
        try:
            power_w = float(power_w)  # На случай, если передали строку
            
            # Проверяем для МВт (от 1 000 000 Вт)
            if power_w >= 1_000_000:
                power_value = power_w / 1_000_000
                # Округляем до 3 знаков после запятой
                return f"{round(power_value, 3)} МВт"
                
            # Проверяем для кВт (от 1 000 Вт)
            elif power_w >= 1_000:
                power_value = power_w / 1_000
                # Округляем до 3 знаков после запятой
                return f"{round(power_value, 3)} кВт"
                
            # Для значений меньше 1 000 Вт
            else:
                # Округляем до 2 знаков после запятой
                return f"{round(power_w, 2)} Вт"
                
        except (ValueError, TypeError):
            # Если возникла ошибка преобразования
            return f"{power_w} Вт"  # Возвращаем как есть с пометкой Вт
        
    def format_weight(self, weight_kg):
        """Форматирует вес с автоматическим выбором единиц измерения"""
        if weight_kg >= 1000:  # Более 1000 кг = 1 т
            return f"{weight_kg / 1000:.3f} т"
        else:
            return f"{weight_kg:.3f} кг"   
        
    def calculate_total_weight_and_volume(self, spec_data):
        """
        Рассчитывает общий вес и объем радиаторов (без учета кронштейнов)
        Возвращает:
            total_weight (float): Суммарный вес в кг
            total_volume (float): Суммарный объем в м³
        """
        total_weight = 0.0
        total_volume = 0.0
        
        # Перебираем все строки в спецификации
        for index, row in spec_data.iterrows():
            # Пропускаем строку "Итого" и кронштейны
            if row["№"] == "Итого" or "Кронштейн" in str(row["Наименование"]):
                continue
            
            # Получаем артикул и количество
            art = str(row["Артикул"]).strip()
            qty = int(row["Кол-во"])
            
            # Ищем радиатор в данных
            for sheet_name, data in self.sheets.items():
                # Проверяем наличие артикула в текущем листе
                product = data[data["Артикул"].str.strip() == art]
                if not product.empty:
                    # Суммируем вес и объем
                    total_weight += float(product.iloc[0]["Вес, кг"]) * qty
                    total_volume += float(product.iloc[0]["Объем, м3"]) * qty
                    break  # Прерываем поиск после нахождения
        
        # Округляем значения как в образце
        return round(total_weight, 1), round(total_volume, 3)

    def on_treeview_motion(self, event, tree):
        """Обработчик движения мыши над Treeview"""
        x = event.x
        y = event.y
        
        # Получаем названия колонок из tree
        columns = [tree.heading(col)["text"] for col in tree["columns"]]
        
        # Определяем столбец под курсором
        column_id = tree.identify_column(x)
        if column_id and y < 25:  # 25px - высота заголовка
            column_index = int(column_id.replace('#', '')) - 1
            if 0 <= column_index < len(columns):
                column_name = columns[column_index]
                if column_name in ["Артикул", "Кол-во"]:
                    self.show_header_tooltip(tree, column_name, event.x_root, event.y_root)
                else:
                    self.hide_header_tooltip()
            else:
                self.hide_header_tooltip()
        else:
            self.hide_header_tooltip()

    def hide_header_tooltip(self):
        """Скрывает текущую подсказку"""
        if self.tooltip:
            self.tooltip.destroy()
            self.tooltip = None               

    def preview_spec(self):
        # Проверяем, существует ли уже окно предпросмотра
        if hasattr(self, '_preview_window') and self._preview_window and self._preview_window.winfo_exists():
            # Если окно существует, делаем его активным и выходим из метода
            self._preview_window.lift()
            self._preview_window.focus_force()
            return
        
        self.tooltip = None
        # Сохраняем значение из текущей активной ячейки (если есть)
        if self.root.focus_get() in self.entries.values():
            for (sheet_name, art), entry in self.entries.items():
                if entry == self.root.focus_get():
                    value = entry.get()
                    if value:
                        self.entry_values[(sheet_name, art)] = value
                    else:
                        self.entry_values.pop((sheet_name, art), None)
                    break
        spec_data = self.prepare_spec_data()
        if spec_data is None or spec_data.empty:
            messagebox.showwarning("Ошибка", "Нет данных для предпросмотра")
            return

        # Создание окна предпросмотра
        preview = tk.Toplevel(self.root)
        preview.title("Предпросмотр спецификации")
        preview.geometry("1400x750+0+0")
        preview.minsize(1200, 600)

        # Главный контейнер
        main_frame = ttk.Frame(preview)
        main_frame.pack(fill="both", expand=True, padx=15, pady=15)

        # Фрейм для таблицы с прокруткой
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill="both", expand=True)

        # Настраиваем стиль для Treeview
        style = ttk.Style()
        
        # Создаем Treeview
        columns = list(spec_data.columns)
        tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show="headings",
            selectmode="extended",
            style='Treeview'
        )
        
        # Добавляем тег для редактируемых ячеек
        tree.tag_configure('editable', foreground='blue')
        
        # Привязываем обработчики событий для редактирования
        tree.bind('<Double-1>', lambda e: self.on_cell_double_click(e, tree, spec_data))
        tree.bind('<Return>', lambda e: self.on_cell_edit_finish(e, tree, spec_data))
        tree.bind('<FocusOut>', lambda e: self.on_cell_edit_finish(e, tree, spec_data))

        # Привязки для подсказок (оставлены без изменений)
        tree.bind("<Motion>", lambda e: self.on_treeview_motion(e, tree))
        tree.bind("<Leave>", lambda e: self.hide_header_tooltip())

        # Настройка столбцов
        col_widths = {
            "№": 50,
            "Артикул": 100,
            "Наименование": 370,
            "Мощность, Вт": 100,
            "Цена, руб (с НДС)": 150,
            "Скидка, %": 80,
            "Цена со скидкой, руб (с НДС)": 230,
            "Кол-во": 120,
            "Сумма, руб (с НДС)": 150
        }

        # Конфигурация заголовков
        for col in columns:
            if col == "Артикул":
                tree.heading(col, text=col, command=lambda: self.copy_articul_column(spec_data))
            elif col == "Кол-во":
                tree.heading(col, text=col, command=lambda: self.copy_quantity_column(spec_data))
            else:
                tree.heading(col, text=col)
                
            tree.column(
                col, 
                width=col_widths.get(col, 100), 
                anchor="center" if col != "Наименование" else "w"
            )

        # Изменяем часть с добавлением данных в Treeview:
        for _, row in spec_data.iterrows():
            # Здесь убираем нули в столбце "Мощность, Вт" для кронштейнов
            power_value = "" if "Кронштейн" in str(row["Наименование"]) else row["Мощность, Вт"]
            
            formatted_row = [
                row["№"],
                str(row["Артикул"]),  # Артикул как строка без форматирования
                row["Наименование"],
                power_value,
                f"{float(row['Цена, руб (с НДС)']):.2f}".replace('.', ','),  # Замена точки на запятую
                f"{float(row['Скидка, %']):.2f}".replace('.', ','),
                f"{float(row['Цена со скидкой, руб (с НДС)']):.2f}".replace('.', ','),
                row["Кол-во"],
                f"{float(row['Сумма, руб (с НДС)']):.2f}".replace('.', ',')
            ]
            tree.insert("", "end", values=formatted_row)

        # Итоговая строка
        total_sum = spec_data["Сумма, руб (с НДС)"].sum()
        total_qty_radiators = sum(spec_data.query("Наименование.str.contains('Радиатор')")["Кол-во"].apply(self.parse_quantity))
        total_qty_brackets = sum(spec_data.query("Наименование.str.contains('Кронштейн')")["Кол-во"].apply(self.parse_quantity))

        total_item = tree.insert("", "end", values=[
            "Итого", "", "", "", "", "", "",
            f"{total_qty_radiators} / {total_qty_brackets}",
            f"{total_sum:.2f}"
        ], tags=("total",))
        tree.tag_configure("total", background="#e0e0e0", font=("Segoe UI", 9, "bold"))

        # Прокрутка
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        # Упаковка элементов
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # Расчет итогов
        total_power = self.calculate_total_power(spec_data)
        total_weight, total_volume = self.calculate_totals(spec_data)

        # Фрейм для итогов
        totals_frame = ttk.Frame(main_frame)
        totals_frame.pack(fill="x", pady=10)

        self.totals_power_label = ttk.Label(totals_frame, 
            text=f"Суммарная мощность: {self.format_power(total_power)}", 
            font=("Segoe UI", 9, "bold"))
        self.totals_power_label.grid(row=0, column=0, padx=15, sticky="w")

        self.totals_weight_label = ttk.Label(totals_frame, 
            text=f"Общий вес: {self.format_weight(total_weight)}", 
            font=("Segoe UI", 9, "bold"))
        self.totals_weight_label.grid(row=0, column=1, padx=15, sticky="w")

        self.totals_volume_label = ttk.Label(totals_frame, 
            text=f"Общий объём: {total_volume:.5f} м³", 
            font=("Segoe UI", 9, "bold"))
        self.totals_volume_label.grid(row=0, column=2, padx=15, sticky="w")

        # Фрейм управления
        control_frame = ttk.Frame(main_frame)
        control_frame.pack(fill="x", pady=10)

        # Левая часть - добавление кронштейнов
        brackets_frame = ttk.Frame(control_frame)
        brackets_frame.pack(side="left", fill="x", expand=True)

        brackets_list = self.get_brackets_list()
        if brackets_list:
            ttk.Label(brackets_frame, text="Добавить кронштейн:").pack(side="left", padx=5)
            
            bracket_combobox = ttk.Combobox(
                brackets_frame,
                values=[b['Наименование'] for b in brackets_list],
                state="readonly",
                width=40
            )
            bracket_combobox.pack(side="left", padx=5)

            # Устанавливаем первый элемент списка как значение по умолчанию
            if brackets_list:
                bracket_combobox.current(0)
            
            ttk.Label(brackets_frame, text="Количество:").pack(side="left", padx=5)
            
            quantity_entry = ttk.Entry(
                brackets_frame,
                width=5,
                validate="key",
                validatecommand=(brackets_frame.register(self.validate_input), '%P')
            )
            quantity_entry.pack(side="left", padx=5)
            
            ttk.Button(
                brackets_frame,
                text="Добавить",
                command=lambda: self.add_bracket_to_spec(
                    bracket_combobox, 
                    quantity_entry, 
                    tree, 
                    spec_data
                )
            ).pack(side="left", padx=5)

        # Правая часть - экспорт
        export_frame = ttk.Frame(control_frame)
        export_frame.pack(side="right", fill="x", padx=10)

        ttk.Button(export_frame, 
            text="Экспорт спецификации", 
            command=lambda: self.generate_spec("excel", tree)
        ).pack(side="left", padx=5)

        ttk.Button(export_frame, 
            text="Экспорт в файл CSV", 
            command=lambda: self.generate_spec("csv", tree)
        ).pack(side="left", padx=5)
        
        # Сохраняем ссылку на spec_data для использования при экспорте
        self._current_spec_data = spec_data

        # Контекстное меню
        self.create_context_menu(tree, spec_data)
        self._preview_window = preview

        # Добавляем обработчик закрытия окна, чтобы обнулить ссылку
        preview.protocol("WM_DELETE_WINDOW", self._close_preview_window)



    def _close_preview_window(self):
        """Закрывает окно предпросмотра и очищает ссылку"""
        if hasattr(self, '_preview_window') and self._preview_window and self._preview_window.winfo_exists():
            self._preview_window.destroy()
        self._preview_window = None    

    def add_bracket_to_spec(self, combobox, entry, tree, spec_data):
        """Добавляет выбранный кронштейн в спецификацию"""
        selected_name = combobox.get()
        qty_str = entry.get()
        
        if not selected_name:
            messagebox.showwarning("Ошибка", "Выберите кронштейн из списка")
            return
            
        if not qty_str:
            messagebox.showwarning("Ошибка", "Введите количество")
            return
        
        try:
            qty = self.parse_quantity(qty_str)
            if qty <= 0:
                raise ValueError("Количество должно быть больше 0")
                
            # Получаем список всех доступных кронштейнов
            brackets_list = self.get_brackets_list()
            
            # Находим выбранный кронштейн по наименованию
            selected_bracket = None
            for bracket in brackets_list:
                if bracket['Наименование'] == selected_name:
                    selected_bracket = bracket
                    break
                    
            if not selected_bracket:
                raise ValueError("Кронштейн не найден в базе данных")
            
            # Ищем информацию о кронштейне в DataFrame
            mask = self.brackets_df['Артикул'] == selected_bracket['Артикул']
            bracket_info = self.brackets_df.loc[mask].iloc[0]
            
            # Рассчитываем цены
            price = float(bracket_info['Цена, руб'])
            discount = float(self.bracket_discount_var.get()) if self.bracket_discount_var.get() else 0.0
            discounted_price = price * (1 - discount / 100)
            total = round(discounted_price * qty, 2)
            
            # После добавления новой строки обновляем сохраненные данные
            if hasattr(self, '_current_spec_data'):
                self._current_spec_data = spec_data

            # Добавляем новую запись в спецификацию
            new_row = {
                "№": len(spec_data) + 1,
                "Артикул": selected_bracket['Артикул'],
                "Наименование": selected_bracket['Наименование'],
                "Мощность, Вт": 0.0,
                "Цена, руб (с НДС)": price,
                "Скидка, %": discount,
                "Цена со скидкой, руб (с НДС)": discounted_price,
                "Кол-во": qty,
                "Сумма, руб (с НДС)": total
            }
            
            spec_data.loc[len(spec_data)] = new_row
            
            # Обновляем Treeview
            self.update_treeview(tree, spec_data)
            
            # Очищаем поля ввода
            combobox.set('')
            entry.delete(0, tk.END)
            
            # Обновляем итоговые значения
            total_power = self.calculate_total_power(spec_data)
            total_weight, total_volume = self.calculate_totals(spec_data)
            
            # Обновляем метки с итогами
            self.totals_power_label.config(text=f"Суммарная мощность: {self.format_power(total_power)}")
            self.totals_weight_label.config(text=f"Общий вес: {self.format_weight(total_weight)}")
            self.totals_volume_label.config(text=f"Общий объём: {total_volume:.5f} м³")
            
        except ValueError as ve:
            messagebox.showerror("Ошибка", str(ve))
        except Exception as e:
            messagebox.showerror("Ошибка", f"Неизвестная ошибка: {str(e)}")

    def show_header_tooltip(self, tree, column_name, x, y):
        """Показывает подсказку над заголовком"""
        if not hasattr(self, 'tooltip') or not self.tooltip:
            self.tooltip = tk.Toplevel(tree)
            self.tooltip.wm_overrideredirect(True)
            self.tooltip_label = ttk.Label(
                self.tooltip,
                text="Нажми для копирования в буфер",
                background="#ffffe0",
                relief="solid",
                padding=5,
                font=("Segoe UI", 9)
            )
            self.tooltip_label.pack()
        
        self.tooltip.wm_geometry(f"+{x - 50}+{y - 40}")
        self.tooltip.deiconify()

    def hide_header_tooltip(self):
        """Скрывает подсказку"""
        if hasattr(self, 'tooltip') and self.tooltip:
            self.tooltip.withdraw()

    def parse_quantity(self, value):
        """
        Преобразует введенное значение в количество радиаторов.
        Обрабатывает целые числа, числа с плавающей точкой и комбинации с плюсами.
        """
        try:
            # Если переданное значение — название столбца (например, "Кол-во"), вернуть 0
            if isinstance(value, str) and value.strip() in ["Кол-во", "№"]:
                return 0
                
            if not value:
                return 0
            
            # Если значение уже число, сразу округляем до ближайшего целого
            if isinstance(value, (int, float)):
                return int(round(float(value)))
        
            value = str(value).strip()
            
            # Удаляем лишние знаки '+' в начале и конце
            while value.startswith('+'):
                value = value[1:]
            while value.endswith('+'):
                value = value[:-1]
            
            # Если осталась пустая строка после очистки, возвращаем 0
            if not value:
                return 0
            
            # Разбиваем строку по знакам '+' и суммируем отдельные части
            parts = value.split('+')
            total = 0
            for part in parts:
                part = part.strip()
                if part:
                    # Преобразуем каждую часть в float, округляем и добавляем к сумме
                    total += int(round(float(part)))
                    
            return total
        except Exception as e:
            print(f"Ошибка преобразования количества: {str(e)}")
            return 0
        
    def reset_fields(self):
        """Сбрасывает все поля и закрывает окно предпросмотра, если оно открыто"""
        self._close_preview_window()
        self.entry_values.clear()
        
        # Очищаем сохраненные данные для спецификации
        if hasattr(self, '_current_spec_data'):
            del self._current_spec_data
            
        # Очищаем данные о соответствии, если они есть
        if hasattr(self, '_correspondence_df'):
            del self._correspondence_df
        
        # Очищаем все поля ввода
        for entry in list(self.entries.values()):  # Создаем копию списка
            try:
                if entry.winfo_exists():  # Проверяем, существует ли виджет
                    entry.delete(0, tk.END)
                    entry.config(bg='white')
            except tk.TclError:
                continue
                
        # Сбрасываем скидки к значениям по умолчанию
        self.radiator_discount_var.set("0")
        self.bracket_discount_var.set("0")
        
        # Сбрасываем тип крепления к значению по умолчанию
        self.bracket_var.set("Настенные кронштейны")
        
        # Убираем флаг добавленных кронштейнов, если он есть
        if hasattr(self, 'preview_brackets_added'):
            del self.preview_brackets_added

    def validate_input(self, P):
        """Проверяет вводимые данные в ячейках. Разрешает цифры и знаки +"""
        if P == "":  # Разрешаем пустую строку
            return True
        return all(char.isdigit() or char == '+' for char in P)

    def validate_discount(self, P):
        if P == "":
            return True
        try:
            float(P)
            return 0 <= float(P) <= 100
        except ValueError:
            return False
    def on_cell_double_click(self, event, tree, spec_data):
        """Обработчик двойного клика для редактирования ячейки"""
        region = tree.identify("region", event.x, event.y)
        if region == "cell":
            column = tree.identify_column(event.x)
            item = tree.identify_row(event.y)
            
            # Разрешаем редактирование только столбца "Кол-во" (8-й столбец)
            if column == "#8":
                x, y, width, height = tree.bbox(item, column)
                
                # Получаем текущее значение
                current_value = tree.item(item, "values")[7]
                
                # Создаем поле ввода
                entry = ttk.Entry(tree, width=10)
                entry.insert(0, current_value)
                entry.place(x=x, y=y, width=width, height=height)
                entry.focus_set()
                
                # Сохраняем ссылки для последующего использования
                self._edit_item = item
                self._edit_column = column
                self._edit_entry = entry
                self._edit_spec_data = spec_data
                
                # Привязываем события
                entry.bind("<Return>", lambda e: self.finish_editing(tree))
                entry.bind("<FocusOut>", lambda e: self.finish_editing(tree))
                entry.bind("<Escape>", lambda e: self.cancel_editing(tree))

    def on_cell_edit_finish(self, event, tree, spec_data):
        """Алиас для finish_editing для привязки событий"""
        self.finish_editing(tree)

    def finish_editing(self, tree):
        """Завершает редактирование и сохраняет значение"""
        if hasattr(self, '_edit_entry'):
            new_value = self._edit_entry.get()
            item = self._edit_item
            
            # Обновляем Treeview
            values = list(tree.item(item, "values"))
            values[7] = new_value
            tree.item(item, values=values)
            
            # Обновляем данные спецификации
            if hasattr(self, '_edit_spec_data'):
                index = int(values[0]) - 1  # Получаем индекс из столбца "№"
                if 0 <= index < len(self._edit_spec_data):
                    # Обновляем количество
                    self._edit_spec_data.at[index, "Кол-во"] = self.parse_quantity(new_value)
                    
                    # Пересчитываем сумму для этой строки
                    price = float(self._edit_spec_data.at[index, "Цена со скидкой, руб (с НДС)"])
                    self._edit_spec_data.at[index, "Сумма, руб (с НДС)"] = price * float(self.parse_quantity(new_value))
                    
                    # Полностью перезагружаем Treeview с обновленными данными
                    self.update_treeview(tree, self._edit_spec_data)
                    
                    # Обновляем итоговые значения под таблицей
                    self.update_footer_totals(self._edit_spec_data)
            
            self._edit_entry.destroy()
            del self._edit_entry
            del self._edit_item
            del self._edit_column
            tree.focus_set()

    def cancel_editing(self, tree):
        """Отменяет редактирование"""
        if hasattr(self, '_edit_entry'):
            self._edit_entry.destroy()
            del self._edit_entry
            del self._edit_item
            del self._edit_column
        tree.focus_set()

    def update_totals(self):
        """Обновляет итоговые значения после редактирования"""
        if hasattr(self, '_current_spec_data'):
            spec_data = self._current_spec_data
            
            # Пересчитываем суммы
            spec_data["Сумма, руб (с НДС)"] = spec_data["Цена со скидкой, руб (с НДС)"] * spec_data["Кол-во"]
            
            # Обновляем Treeview
            if hasattr(self, '_preview_window') and self._preview_window.winfo_exists():
                for widget in self._preview_window.winfo_children():
                    if isinstance(widget, ttk.Frame):
                        for subwidget in widget.winfo_children():
                            if isinstance(subwidget, ttk.Treeview):
                                self.update_treeview(subwidget, spec_data)
                                break
            
            # Обновляем метки с итогами
            total_power = self.calculate_total_power(spec_data)
            total_weight, total_volume = self.calculate_totals(spec_data)
            
            if hasattr(self, 'totals_power_label'):
                self.totals_power_label.config(text=f"Суммарная мощность: {self.format_power(total_power)}")
            if hasattr(self, 'totals_weight_label'):
                self.totals_weight_label.config(text=f"Общий вес: {self.format_weight(total_weight)}")
            if hasattr(self, 'totals_volume_label'):
                self.totals_volume_label.config(text=f"Общий объём: {total_volume:.5f} м³") 

    def update_footer_totals(self, spec_data):
        """Обновляет итоговые значения под таблицей"""
        if hasattr(self, 'totals_power_label') and hasattr(self, 'totals_weight_label') and hasattr(self, 'totals_volume_label'):
            total_power = self.calculate_total_power(spec_data)
            total_weight, total_volume = self.calculate_totals(spec_data)
            
            self.totals_power_label.config(text=f"Суммарная мощность: {self.format_power(total_power)}")
            self.totals_weight_label.config(text=f"Общий вес: {self.format_weight(total_weight)}")
            self.totals_volume_label.config(text=f"Общий объём: {total_volume:.5f} м³")            

if __name__ == "__main__":
    try:
        root = tk.Tk()
        root.withdraw()  # Сначала скрываем окно
        
        # Проверяем, создалось ли окно
        if not root.winfo_exists():
            raise RuntimeError("Не удалось создать главное окно")
            
        app = RadiatorApp(root)
        root.deiconify()  # Показываем окно после инициализации
        root.mainloop()
        
    except Exception as e:
        error_msg = f"Ошибка при запуске приложения: {e}\n\n{traceback.format_exc()}"
        print(error_msg)
        messagebox.showerror("Ошибка", f"Не удалось запустить приложение: {e}")
        sys.exit(1)